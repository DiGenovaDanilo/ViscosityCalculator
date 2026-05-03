"""
app_online.py  —  ViscosityCalculator  (Streamlit web app)
Langhammer et al. (2022) ANN model — two modes:
  1. Viscosity Calculator (anhydrous, multi-sample)
  2. Anhydrous and Hydrous Modelling (Tg, m and viscosity vs H2O)
"""

import sys, os, io, pathlib
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.cm as cm
import matplotlib.patheffects as pe
from scipy.optimize import brentq, minimize
import streamlit as st
import tensorflow as tf
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import urllib.request, zipfile

# ── Paths ─────────────────────────────────────────────────────────────────────
BASE_DIR   = os.path.dirname(os.path.abspath(__file__))
PATH_MODEL = os.path.join(BASE_DIR, "model")
sys.path.insert(0, BASE_DIR)

# ── Page config (MUST be first Streamlit call) ────────────────────────────────
st.set_page_config(page_title="MELVIS — Melt VIScosity", page_icon="🌋", layout="wide")

# ── Auto-download model from Zenodo ──────────────────────────────────────────
ZENODO_URL = "https://zenodo.org/records/19945909/files/model.zip"
if not os.path.exists(PATH_MODEL):
    with st.spinner("Downloading ANN model from Zenodo (first run only, ~50 MB)..."):
        zip_path = os.path.join(BASE_DIR, "model.zip")
        urllib.request.urlretrieve(ZENODO_URL, zip_path)
        with zipfile.ZipFile(zip_path, 'r') as z:
            z.extractall(BASE_DIR)
        os.remove(zip_path)
    st.success("Model downloaded successfully!")

from wt_to_mol_calc import mol_conv
from myega import myega

# ── Constants ─────────────────────────────────────────────────────────────────
OXIDES  = ['SiO2','TiO2','Al2O3','FeO','MnO','MgO','CaO',
           'Na2O','K2O','P2O5','Cr2O3','Fe2O3','H2O']
A_FIXED = -2.9

TEMPLATE_CSV = (
    "Sample,SiO2,TiO2,Al2O3,FeO,MnO,MgO,CaO,Na2O,K2O,P2O5,Cr2O3,Fe2O3,H2O,Reference\n"
    "1,78.1,0.1,12.02,1.6,0.04,0.06,0.9,1.0,5.2,0.02,0,0,0,Unknown\n"
    "2,64.04,0.6,19.0,3.2,0.2,2.3,6.0,4.5,1.6,0,0,0,0,Unknown\n"
    "3,53.0,0.88,20.0,4.0,0,3.18,9.1,4.0,1.0,0.2,0,4.0,0,Unknown\n"
    "4,46.1,2.3,13.0,11.2,0.4,11.6,11.13,2.4,1.1,0.6,0,0,0,Unknown\n"
)

DEFAULT_EXAMPLE_CSV = """Sample,SiO2,TiO2,Al2O3,FeO,MnO,MgO,CaO,Na2O,K2O,P2O5,Cr2O3,Fe2O3,H2O,Reference
JSC-1a,46.6,1.79,16.37,3.79,0.18,8.79,9.96,3.27,0.83,0.71,0.02,7.68,0,"Morrison et al., 2019"
P-MORB,51.0,2.8,14.07,5.98,0.22,5.67,10.47,2.69,0.44,0.0,0.01,6.64,0,"Hofmeister et al., 2016"
NYI-1977,39.57,2.89,14.88,3.76,0.3,4.13,12.16,5.47,5.14,1.47,0.0,10.25,0,"Morrison et al., 2020"
CHW,53.24,1.7,15.75,6.84,0.17,6.43,9.08,2.56,0.68,0.12,0.02,3.41,0,"Sehlke & Whittington, 2016"
Andesite,62.47,0.55,20.03,0.03,0.02,3.22,9.09,3.52,0.93,0.12,0.0,0.0,0,"Richet et al., 1996"
Qtz-latite,67.93,0.95,13.15,3.12,0.1,1.24,2.56,2.69,4.51,0.29,0.0,3.47,0,"Ciocchiatti et al., 1995"
LM,42.27,9.56,8.65,5.67,0.1,5.73,12.71,0.34,0.18,0.06,0.01,14.71,0,"Sehlke & Whittington, 2016"
WPVe,56.4,0.3,21.66,1.15,0.13,0.41,3.28,5.34,9.56,0.0,0.0,1.77,0,"Iacono-Marziano et al., 2007"
Ebu-B,75.32,0.0,8.27,0.93,0.0,0.01,0.25,6.83,3.66,0.0,0.0,4.72,0,"Stabile et al., 2016"
ETN,48.82,1.67,16.96,5.03,0.24,5.53,10.15,3.71,1.85,0.01,0.47,5.58,0,"Di Genova et al., 2014"
NVP-Na,55.78,0.9,15.09,0.96,0.25,13.78,4.35,6.34,0.22,0.01,0.18,2.13,0,"Sehlke & Whittington, 2015"
Vul,53.6,0.03,15.52,4.21,0.16,4.89,8.53,3.67,4.73,0.0,0.0,4.67,0,"Vetere et al., 2007"
FR,56.65,0.81,17.95,3.3,0.17,2.36,5.53,4.56,4.54,0.01,0.45,3.66,0,"Di Genova et al., 2014"
AMS_B1,61.27,0.38,18.38,3.5,0.14,0.74,2.97,4.58,8.04,0.0,0.0,0.0,0,"Romano et al., 2003"
PS-GM,69.23,0.5,9.18,3.71,0.32,0.08,0.6,6.52,4.35,0.04,0.0,5.46,0,"Di Genova et al., 2013"
Bas,43.47,3.87,14.9,6.61,0.2,6.01,10.44,4.36,1.79,1.02,0.0,7.33,0,"Di Fiore et al., 2022"
Ves_G,49.51,0.84,16.5,3.62,0.13,5.13,10.26,2.72,6.54,0.72,0.0,4.02,0,Giordano & Dingwell 2003a
Ves_W,51.78,0.68,18.81,3.08,0.13,2.53,7.38,3.79,7.99,0.4,0.0,3.42,0,Giordano & Dingwell 2003a
A,77.63,0.11,12.73,3.03,0.03,0.06,0.92,4.44,1.62,0.0,0.0,0.0,0,"Di Genova et al., 2017"
"""

# ==============================================================================
# SHARED FUNCTIONS
# ==============================================================================

@st.cache_resource
def load_model():
    return tf.keras.models.load_model(PATH_MODEL)

def redistribute_iron(wt):
    wt = wt.copy()
    feo = wt[3]; fe2o3 = wt[11]
    flag = 'unchanged'
    if feo > 0 and fe2o3 == 0:
        wt[3] = feo/2; wt[11] = feo*1.11/2; flag = 'FeO->split'
    elif fe2o3 > 0 and feo == 0:
        wt[11] = fe2o3/2; wt[3] = fe2o3/1.11/2; flag = 'Fe2O3->split'
    return wt, flag

def normalize_to_100(wt):
    s = wt.sum()
    return wt/s*100.0 if s > 0 else wt

def myega_eq(T, Tg, m, A=A_FIXED):
    return A + (Tg/T)*(12-A)*np.exp(((m/(12-A))-1)*((Tg/T)-1))

def visc_calc_fast(inp, si, model):
    eta_goal = ([0.0,0.5,1.0,1.5,2.0,9.5,10,10.5,11.0,11.5] if si<=60
                else [2,2.5,3,3.5,4,4.5,9.5,10,10.5,11.0,11.5])
    t_max = 2023.0
    t_top = (3000.0/t_max-0.602847523)/np.sqrt(0.031535353)
    t_bot = (300.0 /t_max-0.602847523)/np.sqrt(0.031535353)
    t_mid = (t_top+t_bot)/2
    comp = np.c_[t_mid,inp[0],inp[1],inp[2],inp[3],
                 inp[4],inp[5],inp[6],inp[7],inp[8],
                 inp[9],inp[11],inp[12],inp[13],inp[14]]
    eta_mid = model(comp)
    t_goal, eta = [], []
    for goal in eta_goal:
        t_top2=t_top; t_bot2=t_bot; t_mid2=t_mid; eta_mid2=eta_mid
        err=eta_mid2-goal
        while np.absolute(err)>1e-3:
            if eta_mid2<goal: t_top2=t_mid2
            else:             t_bot2=t_mid2
            t_mid2=     (t_bot2+t_top2)/2
            comp[0][0]=t_mid2
            eta_mid2=model(comp)
            eta_mid2=np.concatenate(eta_mid2)
            err=eta_mid2-goal
        t_goal.append((t_mid2*np.sqrt(0.031535353)+0.602847523)*t_max)
        eta.append(float(np.array(eta_mid2).flatten()[0]))
    return np.array(t_goal,dtype=float), np.array(eta,dtype=float)

def write_sheet(ws, df_in, hdr_color='1F4E79'):
    hdr_font = Font(name='Arial',bold=True,color='FFFFFF',size=10)
    hdr_fill = PatternFill('solid',start_color=hdr_color)
    alt_fill = PatternFill('solid',start_color='D6E4F0')
    nrm_fill = PatternFill('solid',start_color='FFFFFF')
    thin     = Side(style='thin',color='AAAAAA')
    brd      = Border(left=thin,right=thin,top=thin,bottom=thin)
    ctr      = Alignment(horizontal='center',vertical='center')
    cols = list(df_in.columns)
    for c,col in enumerate(cols,1):
        cell=ws.cell(row=1,column=c,value=col)
        cell.font=hdr_font; cell.fill=hdr_fill
        cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        cell.border=brd
    ws.row_dimensions[1].height=25
    for r,row_data in enumerate(df_in.itertuples(index=False),2):
        fill=alt_fill if r%2==0 else nrm_fill
        for c,val in enumerate(row_data,1):
            cell=ws.cell(row=r,column=c,value=val if pd.notna(val) else '')
            cell.border=brd; cell.fill=fill; cell.alignment=ctr
            if isinstance(val,float): cell.number_format='0.000'
    for c,col in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(c)].width=max(len(str(col)),8)+3
    ws.freeze_panes='B2'

# ── Hydrous modelling helpers ─────────────────────────────────────────────────

def get_Tg_m(wt_anhydrous, h2o_wt, model):
    wt = wt_anhydrous.copy()
    wt[12] = h2o_wt
    wt, _ = redistribute_iron(wt)
    wt     = normalize_to_100(wt)
    normalised, _, _ = mol_conv(wt)
    t_synth, eta_synth = visc_calc_fast(normalised, wt[0], model)
    params, _, _, _ = myega(t_synth, eta_synth, np.array([1000.0]))
    return params[0], params[1]   # Tg (K), m

def wt_to_mol_h2o(h2o_wt_pct, wt_anhydrous):
    wt = wt_anhydrous.copy(); wt[12] = h2o_wt_pct
    wt = normalize_to_100(wt)
    _, _, mol_per = mol_conv(wt)
    return mol_per[12]

def tg_model(x_h2o_mol, b, c, d, Tg_d, Tg_H2O=136.0):
    denom = b*(100.0-x_h2o_mol) + x_h2o_mol
    if denom == 0: return Tg_d
    w1 = x_h2o_mol / denom
    w2 = b*(100.0-x_h2o_mol) / denom
    return (w1*Tg_H2O + w2*Tg_d
            + c*w1*w2*(Tg_d-Tg_H2O)
            + d*w1*w2**2*(Tg_d-Tg_H2O))

def m_from_tg(Tg, Tg_d, m_d, A=A_FIXED):
    return m_d + (12-A)*np.log(Tg/Tg_d)

def fit_tg(x_mol_arr, Tg_arr, Tg_d):
    def rmse_tg(params):
        b, c, d = params
        if b <= 0: return 1e10
        try:
            pred = np.array([tg_model(x, b, c, d, Tg_d) for x in x_mol_arr])
            if not np.all(np.isfinite(pred)): return 1e10
            return np.sqrt(np.mean((pred-Tg_arr)**2))
        except: return 1e10
    best_rmse = np.inf
    best_p    = np.array([0.1, 1.5, -2.0])
    for b0 in [0.05,0.1,0.15,0.2,0.3,0.5,1.0,2.0]:
        for c0 in [0.5,1.0,1.5,2.0,3.0]:
            for d0 in [-3.0,-2.0,-1.5,-1.0,-0.5]:
                try:
                    res = minimize(rmse_tg,[b0,c0,d0],method='Nelder-Mead',
                                   options={'maxiter':20000,'xatol':1e-8,'fatol':1e-8})
                    if res.fun < best_rmse and res.x[0]>0:
                        best_rmse=res.fun; best_p=res.x
                except: pass
    return best_p, best_rmse

def make_visc_sheet_hydrous(wb, sheet_name, results, m_func, tg_model_func, hdr_color):
    ws = wb.create_sheet(sheet_name)
    BLOCK=3; GAP=1; T_STEP=25
    for s_idx, r in enumerate(results):
        col_start = 1+s_idx*(BLOCK+GAP)
        Tg_f = tg_model_func(r['h2o_mol'])
        m_v  = m_func(r['h2o_mol'])
        ws.merge_cells(start_row=1,start_column=col_start,end_row=1,end_column=col_start+BLOCK-1)
        hdr = ws.cell(row=1,column=col_start,
            value='{:.1f} wt% H2O | Tg={:.1f}C | m={:.2f}'.format(r['h2o_wt'],Tg_f-273.15,m_v))
        thin=Side(style='thin',color='AAAAAA')
        brd=Border(left=thin,right=thin,top=thin,bottom=thin)
        ctr=Alignment(horizontal='center',vertical='center')
        hdr.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
        hdr.fill=PatternFill('solid',start_color=hdr_color)
        hdr.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
        hdr.border=brd
        for ci,label in enumerate(['T (C)','T (K)','log10(visc/Pa.s)'],0):
            cell=ws.cell(row=2,column=col_start+ci,value=label)
            cell.font=Font(name='Arial',bold=True,color='FFFFFF',size=9)
            cell.fill=PatternFill('solid',start_color='37474F')
            cell.alignment=ctr; cell.border=brd
        ws.row_dimensions[2].height=25
        try: T_max=brentq(myega_eq,Tg_f,5000.0,args=(Tg_f,m_v))
        except: T_max=3000.0
        T_first=int((Tg_f-273.15)//T_STEP)*T_STEP
        T_list=list(range(T_first,int(T_max)+T_STEP,T_STEP))
        alt_f=PatternFill('solid',start_color='D6E4F0')
        wht_f=PatternFill('solid',start_color='FFFFFF')
        for ri,tc in enumerate(T_list):
            row_n=3+ri; tk_val=tc+273.15
            visc_v=round(float(myega_eq(tk_val,Tg_f,m_v)),4)
            fill=alt_f if ri%2==0 else wht_f
            for ci,val in enumerate([float(tc),round(tk_val,2),visc_v],0):
                cell=ws.cell(row=row_n,column=col_start+ci,value=val)
                cell.alignment=ctr; cell.border=brd; cell.fill=fill
                if isinstance(val,float): cell.number_format='0.000'
        for ci in range(BLOCK):
            ws.column_dimensions[get_column_letter(col_start+ci)].width=14
        if s_idx<len(results)-1:
            ws.column_dimensions[get_column_letter(col_start+BLOCK)].width=2
    ws.freeze_panes='A3'

# ==============================================================================
# SESSION STATE
# ==============================================================================
for key in ['calc_done','all_curves','rows_recalc','tg_m_dict','skipped',
            'fig','df_input','rows_specific',
            'hyd_done','hyd_results','hyd_fig','hyd_buf_excel',
            'hyd_buf_fig','hyd_meta','visc_h2o','wt_dry','non_mono','df_h_selected']:
    if key not in st.session_state:
        st.session_state[key] = None
if st.session_state['calc_done'] is None:
    st.session_state['calc_done'] = False
if st.session_state['hyd_done'] is None:
    st.session_state['hyd_done'] = False

# ==============================================================================
# SIDEBAR
# ==============================================================================
with st.sidebar:
    st.markdown("## 🌋 MELVIS")
    st.caption("**MEL**t **VIS**cosity — volcanic melt viscosity platform by [GLASS laboratory](https://www.danilodigenova.org/glass-laboratory/)")
    st.divider()
    mode = st.radio(
        "**Select mode:**",
        [
            "🏠 Home",
            "🔥 Viscosity Calculator",
            "💧 Anhydrous and Hydrous Modelling",
            "🌋 Specific Composition Models",
        ],
        index=0
    )
    st.divider()
    if mode == "🌋 Specific Composition Models":
        pass  # description shown in main area
    elif mode == "🔥 Viscosity Calculator":
        st.markdown("""
**Input format**

CSV with a **Sample** column and oxide columns in wt%
(missing ones set to 0 automatically):

`SiO2, TiO2, Al2O3, FeO, MnO, MgO, CaO, Na2O, K2O, P2O5, Cr2O3, Fe2O3, H2O`

Optional: `Reference` column.

**Iron redistribution:**
- Only FeO → split 50/50 FeO / Fe₂O₃
- Only Fe₂O₃ → split 50/50 Fe₂O₃ / FeO
- Both present → no change
        """)
        st.divider()
        st.markdown("""
**References:**  
Langhammer et al. (2022), *GGG*  
[doi:10.1029/2022GC010673](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2022GC010673)

Langhammer et al. (2021), *GGG*  
[doi:10.1029/2021GC009918](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918)
        """)
    elif mode == "💧 Anhydrous and Hydrous Modelling":
        st.markdown("""
**Input format**

CSV with a **single anhydrous composition** (H₂O = 0).

Same oxide columns as Mode 1.

The tool fits Tg(H₂O) with Eq. 9-10 and computes
fragility m via Eq. 12 (Langhammer et al. 2021).
        """)
        st.divider()
        st.markdown("""
**References:**  
Langhammer et al. (2022), *GGG*  
[doi:10.1029/2022GC010673](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2022GC010673)

Langhammer et al. (2021), *GGG*  
[doi:10.1029/2021GC009918](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918)
        """)
    elif mode == "🌋 Specific Composition Models":
        st.markdown("📖 References shown below each model.")
    if mode != "🌋 Specific Composition Models":
        st.divider()
    st.markdown("**Questions?** ✉️ [danilo.digenova@cnr.it](mailto:danilo.digenova@cnr.it)")

# ==============================================================================
# ==============================================================================
# HOME — LANDING PAGE
# ==============================================================================
if mode == "🏠 Home":

    st.markdown("<div style='text-align:center'><span style='font-size:3em'>🌋</span></div>", unsafe_allow_html=True)
    st.title("MELVIS — Melt VIScosity")
    st.markdown("*Volcanic melt viscosity platform developed by the [GLASS laboratory](https://www.danilodigenova.org/glass-laboratory/), CNR-ISSMC, Rome, Italy*")
    st.divider()

    _c1,_c2,_c3 = st.columns(3)
    with _c1:
        st.markdown("""
**🔥 Viscosity Calculator**  
Calculate viscosity for any silicate melt composition using the Langhammer et al. (2022) ANN.  
Upload a CSV, visualise MYEGA curves, TAS diagram, and download Excel outputs.
        """)
    with _c2:
        st.markdown("""
**💧 Anhydrous and Hydrous Modelling**  
Model how viscosity evolves with dissolved H₂O for a single composition.  
Fits Tg(H₂O) and fragility m(H₂O) and compares fragility models.
        """)
    with _c3:
        st.markdown("""
**🌋 Specific Composition Models**  
MYEGA models calibrated on specific volcanic compositions:  
Stromboli basalt, Peridotite, HPG8+Na, Anhydrous andesite, Vesuvio phonotephrite (472 CE).
        """)

    st.divider()
    st.subheader("About GLASS laboratory")

    # Logos row — loaded client-side via HTML img tags (bypasses Streamlit server proxy)
    st.markdown("""
<div style="display:flex; align-items:center; gap:40px; margin-bottom:20px; flex-wrap:wrap;">
  <div style="text-align:center">
    <img src="https://raw.githubusercontent.com/consiglionazionaledellericerche/cool-jconon/HEAD/cool-jconon-webapp-resources/src/main/resources/META-INF/img/logo-quadrato-en.png"
         height="160" style="object-fit:contain"><br>
    <small style="color:gray">CNR</small>
  </div>
  <div style="text-align:center">
    <img src="https://www.cnr.it/it/istituto/073/logo"
         height="160" style="object-fit:contain"><br>
    <small style="color:gray">CNR-ISSMC</small>
  </div>
  <div style="text-align:center">
    <img src="https://erc.europa.eu/sites/default/files/inline-images/HE%20logo.png"
         height="160" style="object-fit:contain"><br>
    <small style="color:gray">ERC</small>
  </div>
</div>
""", unsafe_allow_html=True)

    _g1,_g2 = st.columns([2,1])
    with _g1:
        st.markdown("""
MELVIS is developed by the [**GLASS laboratory**](https://www.danilodigenova.org/glass-laboratory/) 
(**G**ateway **L**aboratory of **A**morphous and **S**tructured **S**olids and Melts),
established in **2023 in Rome** with funding from the **European Research Council (ERC)
Consolidator Grant NANOVOLC** ([grant 101044772](https://cordis.europa.eu/project/id/101044772)).

GLASS operates within the **National Research Council of Italy ([CNR](https://www.cnr.it/en))**
at the Institute of Science, Technology and Sustainability for Ceramics
([ISSMC](https://www.issmc.cnr.it/en/)).

**Research focus:** physicochemical behaviour of silicate melts, magmas, and glasses,
with emphasis on volcanic processes, eruption dynamics, and advanced materials design.

**Principal Investigator:** [Danilo Di Genova](https://www.danilodigenova.org/) (Research Director at CNR-ISSMC)  
📧 [danilo.digenova@cnr.it](mailto:danilo.digenova@cnr.it)
        """)
    with _g2:
        st.markdown("""
**🔬 Key research areas:**
- Silicate melt rheology
- Volcanic eruption dynamics
- Glass transition & fragility
- Magma degassing & water
- Advanced ceramics & glasses
- Sustainable materials
        """)

    st.divider()
    st.subheader("How to cite MELVIS")
    st.markdown("""
Please cite the relevant model references shown in each module and acknowledge:

> *MELVIS — GLASS laboratory, CNR-ISSMC, Rome, Italy.
> ERC Consolidator Grant NANOVOLC (grant 101044772).*
    """)
    st.info("👈 Select a mode from the sidebar to get started.")

# ==============================================================================
# MODE 1 — VISCOSITY CALCULATOR
# ==============================================================================
if mode == "🔥 Viscosity Calculator":

    st.title("🔥 Viscosity Calculator")
    st.markdown("Calculate silicate melt viscosity from composition using the **Langhammer et al. (2022)** ANN model. Upload a CSV with multiple samples and download MYEGA curves.")

    input_method_1 = st.radio("**How do you want to provide the composition?**",
                              ["📂 Upload CSV file", "⌨️ Type composition manually"],
                              horizontal=True, key="mode1_input_method")

    df = None

    if input_method_1 == "📂 Upload CSV file":
        st.download_button("📄 Download CSV template", data=TEMPLATE_CSV,
                           file_name="Example_compositions.csv", mime="text/csv")
        uploaded = st.file_uploader("Upload your CSV file", type=["csv"])
        if uploaded is None:
            # Default: load bundled Example_compositions.csv
            _default_csv = pathlib.Path(BASE_DIR) / "Example_compositions.csv"
            if _default_csv.exists():
                df = pd.read_csv(_default_csv).dropna(how='all').reset_index(drop=True)
            else:
                import io as _sio
                df = pd.read_csv(_sio.StringIO(DEFAULT_EXAMPLE_CSV)).dropna(how='all').reset_index(drop=True)
            st.info("📋 Using bundled **Example_compositions.csv** — upload your own CSV file above to replace it.")
        else:
            df = pd.read_csv(uploaded).dropna(how='all').reset_index(drop=True)
        for ox in OXIDES:
            if ox not in df.columns: df[ox] = 0.0
        df[OXIDES] = df[OXIDES].fillna(0.0)
        st.success(f"✅ File loaded: **{len(df)} samples** found.")
        with st.expander("Preview input data"): st.dataframe(df)

    else:  # Manual input
        st.markdown("**Enter your composition in wt% (one sample):**")
        m1_name = st.text_input("Sample name:", value="My_sample", key="m1_sname")
        OXIDES_NO_H2O = [o for o in OXIDES if o != "H2O"]
        cols_m1 = st.columns(4)
        defaults_m1 = {
            "SiO2": 48.05, "TiO2": 0.76, "Al2O3": 17.69, "FeO": 6.08,
            "MnO": 0.14, "MgO": 3.32, "CaO": 9.31, "Na2O": 3.45,
            "K2O": 7.55, "P2O5": 0.46, "Cr2O3": 0.0, "Fe2O3": 0.0,
        }
        manual_vals_m1 = {}
        for i, ox in enumerate(OXIDES_NO_H2O):
            with cols_m1[i % 4]:
                manual_vals_m1[ox] = st.number_input(
                    ox, min_value=0.0, max_value=100.0,
                    value=float(defaults_m1.get(ox, 0.0)),
                    step=0.01, format="%.3f", key=f"m1_manual_{ox}")
        manual_vals_m1["H2O"] = st.number_input(
            "H2O", min_value=0.0, max_value=20.0, value=0.0,
            step=0.01, format="%.3f", key="m1_manual_H2O")
        total_m1 = sum(manual_vals_m1[o] for o in OXIDES)
        st.caption(f"Sum of oxides: **{total_m1:.2f} wt%** — will be normalised to 100%")
        df = pd.DataFrame([{"Sample": m1_name, **manual_vals_m1}])
        st.success(f"✅ Composition ready: **{m1_name}**")
        with st.expander("Preview composition"): st.dataframe(df)

    if st.button("▶️ Calculate viscosity", type="primary"):
        model = load_model()
        all_curves=[]; rows_recalc=[]; skipped=[]; tg_m_dict={}
        progress=st.progress(0,text="Calculating...")
        fig,ax=plt.subplots(figsize=(12,7))
        colors=plt.cm.nipy_spectral(np.linspace(0,1,len(df)))

        for idx,row in df.iterrows():
            sname=row['Sample']
            wt_orig=np.array([row[o] for o in OXIDES],dtype=float)
            wt_fe,fe_flag=redistribute_iron(wt_orig)
            wt_final=normalize_to_100(wt_fe)
            rec={'Sample':sname}
            for i,ox in enumerate(OXIDES): rec[ox]=round(wt_final[i],4)
            rec['SUM']=round(wt_final.sum(),4); rec['Fe_treatment']=fe_flag
            if 'Reference' in df.columns: rec['Reference']=row['Reference']
            rows_recalc.append(rec)
            try:
                normalised,_,_=mol_conv(wt_final)
                t_synth,eta_synth=visc_calc_fast(normalised,wt_final[0],model)
                if not np.all(np.isfinite(t_synth)) or not np.all(np.isfinite(eta_synth)):
                    raise ValueError("ANN returned NaN")
                param,_,_,_=myega(t_synth,eta_synth,np.array([1000.0]))
                Tg=param[0]; m=param[1]
                tg_m_dict[sname]=(Tg,m)
                try: T_max=brentq(myega_eq,Tg,5000.0,args=(Tg,m))
                except: T_max=3000.0
                T_array=np.arange(Tg,T_max+50,50)
                visc_array=myega_eq(T_array,Tg,m)
                ax.plot(T_array-273.15,visc_array,color=colors[idx],linewidth=1.5,label=sname)
                for i,(T,v) in enumerate(zip(T_array,visc_array)):
                    all_curves.append({
                        'Sample':   sname if i==0 else '',
                        'Tg_K':    round(Tg,1)        if i==0 else '',
                        'Tg_C':    round(Tg-273.15,1) if i==0 else '',
                        'm':       round(m,2)          if i==0 else '',
                        'T_K':     round(T,1),
                        'T_C':     round(T-273.15,1),
                        'log10_visc': round(float(v),3),
                    })
                all_curves.append({k:'' for k in ['Sample','Tg_K','Tg_C','m','T_K','T_C','log10_visc']})
            except Exception as e:
                skipped.append({'Sample':sname,'Error':str(e)})
            progress.progress((idx+1)/len(df), text=f"Processing {idx+1}/{len(df)}: {sname}")

        progress.empty()
        ax.set_xlabel('Temperature (°C)',fontsize=13)
        ax.set_ylabel('log₁₀(Viscosity / Pa·s)',fontsize=13)
        ax.legend(fontsize=6,loc='upper right',ncol=3,framealpha=0.7,handlelength=1.5)
        ax.grid(True,linestyle='--',alpha=0.5)
        plt.tight_layout()
        st.session_state.update({'calc_done':True,'all_curves':all_curves,
            'rows_recalc':rows_recalc,'tg_m_dict':tg_m_dict,
            'skipped':skipped,'fig':fig,'df_input':df,'rows_specific':[]})

    if st.session_state['calc_done']:
        all_curves=st.session_state['all_curves']
        rows_recalc=st.session_state['rows_recalc']
        tg_m_dict=st.session_state['tg_m_dict']
        skipped=st.session_state['skipped']
        fig=st.session_state['fig']
        df_input=st.session_state['df_input']

        st.subheader("📈 Viscosity curves")
        st.pyplot(fig)

        # ── TAS diagram ──────────────────────────────────────────────────────────
        st.subheader("🗺️ TAS diagram with viscosity colormap")

        tg_m_dict_tas  = st.session_state.get('tg_m_dict')
        df_input_tas   = st.session_state.get('df_input')

        if tg_m_dict_tas and df_input_tas is not None:
            tg_m_dict = tg_m_dict_tas
            df_input  = df_input_tas

            # ── Sub-classification: Na2O-K2O criterion (Le Maitre 2002) ───────
            # For fields S1, S2, S3 (Trachy-basalt, Bas.Trachy-andesite, Trachy-andesite)
            # Na2O - 2 >= K2O → sodic series; Na2O - 2 < K2O → potassic series
            SUBCLASS_MAP = {
                'S1': ('Hawaiite',              'Potassic trachybasalt'),
                'S2': ('Mugearite',             'Shoshonite'),
                'S3': ('Benmoreite',            'Latite'),
            }

            import warnings; warnings.filterwarnings('ignore')
            from pyrolite.util.classification import TAS as TASclf
            _tas_clf = TASclf()

            # Build TAS data from input
            tas_data = []
            for _, row in df_input.iterrows():
                sname = row['Sample']
                if sname in tg_m_dict:
                    Tg_s, m_s = tg_m_dict[sname]
                    sio2  = float(row['SiO2'])
                    na2o  = float(row['Na2O'])
                    k2o   = float(row['K2O'])
                    tas_v = na2o + k2o
                    # Get TAS field ID
                    try:
                        _pred_df = pd.DataFrame({'SiO2': [sio2], 'Na2O + K2O': [tas_v]})
                        field_id = str(_tas_clf.predict(_pred_df).iloc[0])
                    except Exception:
                        field_id = ''
                    # Apply sub-classification
                    if field_id in SUBCLASS_MAP:
                        sodic, potassic = SUBCLASS_MAP[field_id]
                        subclass = sodic if (na2o - 2) >= k2o else potassic
                    else:
                        subclass = ''
                    tas_data.append({
                        'Sample':   sname,
                        'SiO2':     sio2,
                        'TAS':      tas_v,
                        'Na2O':     na2o,
                        'K2O':      k2o,
                        'Tg_C':     round(Tg_s - 273.15, 1),
                        'm':        round(m_s, 2),
                        'field_id': field_id,
                        'subclass': subclass,
                    })
            df_tas = pd.DataFrame(tas_data)

            tas_color = st.selectbox(
                "Color samples by:",
                ["Tg (°C)", "Fragility m", "log₁₀η at custom T (°C)"],
                key="tas_color_sel")

            if tas_color == "log₁₀η at custom T (°C)":
                T_tas = st.number_input("Temperature for η (°C):",
                                        min_value=500.0, max_value=1800.0,
                                        value=1200.0, step=50.0, key="tas_T")
                df_tas['color_val'] = [
                    float(myega_eq(T_tas+273.15, tg_m_dict[r['Sample']][0],
                                   tg_m_dict[r['Sample']][1]))
                    for _, r in df_tas.iterrows()
                ]
                cbar_label = f'log₁₀(η / Pa·s) at {T_tas:.0f} °C'
            elif tas_color == "Tg (°C)":
                df_tas['color_val'] = df_tas['Tg_C']
                cbar_label = 'Tg (°C)'
            else:
                df_tas['color_val'] = df_tas['m']
                cbar_label = 'Fragility index m'

            # ── Draw TAS diagram ──────────────────────────────────────────────────
            fig_tas, ax_tas = plt.subplots(figsize=(10, 7))

            # Use pyrolite for correct TAS (Middlemost 1994)
            import warnings; warnings.filterwarnings('ignore')
            from pyrolite.util.classification import TAS as TASclf
            import matplotlib.patheffects as pe

            tas_clf = TASclf()
            tas_clf.add_to_axes(ax=ax_tas, add_labels=True,
                                which_labels='volcanic',
                                facecolor='#F0F0F0', edgecolor='#888888',
                                linewidth=1.0, alpha=0.9)
            for txt in ax_tas.texts:
                txt.set_fontsize(7)
                txt.set_color('#555555')
                txt.set_path_effects([pe.withStroke(linewidth=2, foreground='white')])

            # Scatter samples
            sc = ax_tas.scatter(df_tas['SiO2'], df_tas['TAS'],
                                c=df_tas['color_val'],
                                cmap='plasma_r', s=120, zorder=5,
                                edgecolors='black', linewidths=0.8)
            plt.colorbar(sc, ax=ax_tas, label=cbar_label, shrink=0.8)

            # Label samples (add sub-classification if available)
            for _, row in df_tas.iterrows():
                label = row['Sample']
                if row.get('subclass', ''):
                    label += f"\n({row['subclass']})"
                ax_tas.annotate(label,
                               xy=(row['SiO2'], row['TAS']),
                               xytext=(4, 4), textcoords='offset points',
                               fontsize=7, zorder=6, linespacing=1.3,
                               path_effects=[pe.withStroke(linewidth=2,
                                                           foreground='white')])

            ax_tas.set_xlim(37, 80)
            ax_tas.set_ylim(0, 17)
            ax_tas.set_xlabel('SiO₂ (wt%)', fontsize=12)
            ax_tas.set_ylabel('Na₂O + K₂O (wt%)', fontsize=12)
            ax_tas.set_title('TAS diagram — ' + cbar_label, fontsize=12, fontweight='bold')
            ax_tas.grid(True, linestyle='--', alpha=0.3, zorder=0)
            plt.tight_layout()

            st.pyplot(fig_tas)

            buf_tas = io.BytesIO()
            fig_tas.savefig(buf_tas, format='png', dpi=200, bbox_inches='tight')
            buf_tas.seek(0)
            plt.close(fig_tas)

            st.download_button("⬇️ Download TAS plot (PNG)",
                               data=buf_tas,
                               file_name="TAS_diagram.png",
                               mime="image/png",
                               key="dl_tas")
            st.caption("📖 TAS after Middlemost (1994), *Earth-Science Reviews* 37, 215-224. Fields from Le Bas et al. (1992), *Mineralogy and Petrology* 46, 1-22. Sub-classification (Hawaiite/Mugearite/Benmoreite vs Potassic trachybasalt/Shoshonite/Latite) from Le Maitre et al. (2002), *Igneous Rocks: A Classification and Glossary of Terms*. Implemented via [pyrolite](https://pyrolite.readthedocs.io/).")

    st.subheader("🌡️ Viscosity at specific temperatures (optional)")
    do_specific = st.checkbox("Calculate viscosity at specific temperatures?")
    rows_specific = st.session_state.get('rows_specific', []) or []

    if do_specific:
        sample_names = list(tg_m_dict.keys())
        selected = st.multiselect("Select samples:", sample_names, default=sample_names)
        if selected:
            st.markdown("**Enter temperatures (°C) for each sample:**")
            temps_per_sample = {}
            for sname in selected:
                t_input = st.text_input(
                    f"Temperatures for **{sname}** (comma-separated):",
                    placeholder="e.g. 800, 1000, 1200, 1400",
                    key=f"temps_{sname}")
                temps_per_sample[sname] = t_input
            if st.button("✅ Compute specific temperatures"):
                rows_specific = []
                for sname, t_input in temps_per_sample.items():
                    if not t_input.strip(): continue
                    try:
                        temps = [float(t.strip()) for t in t_input.split(',') if t.strip()]
                        Tg, m = tg_m_dict[sname]
                        for tc in sorted(temps):
                            rows_specific.append({
                                'Sample': sname, 'T_C': round(tc, 2),
                                'T_K': round(tc+273.15, 2),
                                'log10_visc': round(float(myega_eq(tc+273.15, Tg, m)), 4)})
                    except ValueError:
                        st.error(f"Invalid temperatures for {sname}.")
                st.session_state['rows_specific'] = rows_specific
        if st.session_state.get('rows_specific'):
            st.dataframe(pd.DataFrame(st.session_state['rows_specific']))

        # ── Build Excel ───────────────────────────────────────────────────────
        rows_specific=st.session_state.get('rows_specific',[]) or []
        wb=Workbook()
        ws1=wb.active; ws1.title='Viscosity_Curves'
        write_sheet(ws1,pd.DataFrame(all_curves),hdr_color='1F4E79')
        if rows_specific:
            ws2=wb.create_sheet('Viscosity_at_T')
            write_sheet(ws2,pd.DataFrame(rows_specific),hdr_color='4A235A')

            # MYEGA_Calculator sheet
        ws_calc=wb.create_sheet('MYEGA_Calculator')
        T_STEP_C=25; T_END_C=1600
        inp_fill=PatternFill('solid',start_color='FFF9C4')
        res_fill=PatternFill('solid',start_color='E3F2FD')
        wht_fill=PatternFill('solid',start_color='FFFFFF')
        drk_fill=PatternFill('solid',start_color='37474F')
        thin_c=Side(style='thin',color='AAAAAA')
        brd_c=Border(left=thin_c,right=thin_c,top=thin_c,bottom=thin_c)
        ctr_c=Alignment(horizontal='center',vertical='center')
        BCOLS=3; GCOLS=1
        for s_idx,sname in enumerate(tg_m_dict):
            Tg_val,m_val=tg_m_dict[sname]
            cs=1+s_idx*(BCOLS+GCOLS)
            ws_calc.merge_cells(start_row=1,start_column=cs,end_row=1,end_column=cs+BCOLS-1)
            ch=ws_calc.cell(row=1,column=cs,value=sname)
            ch.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
            ch.fill=PatternFill('solid',start_color='1B5E20')
            ch.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            ch.border=brd_c; ws_calc.row_dimensions[1].height=22
            ws_calc.cell(row=2,column=cs,value='Tg (K)').font=Font(name='Arial',bold=True,size=9)
            tc2=ws_calc.cell(row=2,column=cs+1,value=round(Tg_val,1))
            tc2.fill=inp_fill; tc2.border=brd_c; tc2.alignment=ctr_c; tc2.number_format='0.0'
            ws_calc.cell(row=3,column=cs,value='m').font=Font(name='Arial',bold=True,size=9)
            mc2=ws_calc.cell(row=3,column=cs+1,value=round(m_val,2))
            mc2.fill=inp_fill; mc2.border=brd_c; mc2.alignment=ctr_c; mc2.number_format='0.00'
            tga=f"${get_column_letter(cs+1)}$2"; ma=f"${get_column_letter(cs+1)}$3"
            DR=5
            for ci,label in enumerate(['T (°C)','T (K)','log10(visc / Pa·s)'],0):
                cell=ws_calc.cell(row=DR-1,column=cs+ci,value=label)
                cell.font=Font(name='Arial',bold=True,color='FFFFFF',size=9)
                cell.fill=drk_fill; cell.alignment=ctr_c; cell.border=brd_c
            ws_calc.row_dimensions[DR-1].height=28
            Tg_C=Tg_val-273.15; T_first=int(Tg_C//T_STEP_C)*T_STEP_C
            T_values=list(range(T_first,T_END_C+T_STEP_C,T_STEP_C))
            for r_idx,tc in enumerate(T_values):
                dr=DR+r_idx; fill=res_fill if r_idx%2==0 else wht_fill
                tcc=get_column_letter(cs)
                ct=ws_calc.cell(row=dr,column=cs,value=tc)
                ct.border=brd_c; ct.alignment=ctr_c; ct.fill=fill
                ctk=ws_calc.cell(row=dr,column=cs+1,value=f"={tcc}{dr}+273.15")
                ctk.border=brd_c; ctk.alignment=ctr_c; ctk.number_format='0.00'; ctk.fill=fill
                tka=f"{get_column_letter(cs+1)}{dr}"
                formula=f"=-2.9+({tga}/{tka})*(12-(-2.9))*EXP(({ma}/(12-(-2.9))-1)*({tga}/{tka}-1))"
                cv=ws_calc.cell(row=dr,column=cs+2,value=formula)
                cv.border=brd_c; cv.alignment=ctr_c; cv.number_format='0.000'; cv.fill=fill
                ws_calc.column_dimensions[get_column_letter(cs)].width=9
                ws_calc.column_dimensions[get_column_letter(cs+1)].width=9
                ws_calc.column_dimensions[get_column_letter(cs+2)].width=16
                if s_idx<len(tg_m_dict)-1:
                    ws_calc.column_dimensions[get_column_letter(cs+BCOLS)].width=2
        ws_calc.freeze_panes='A5'

            # Chemistry sheet
        wb_chem=Workbook()
        ws_in=wb_chem.active; ws_in.title='Input_Chemistry'
        cols_orig=['Sample']+OXIDES
        if 'Reference' in df_input.columns: cols_orig.append('Reference')
        df_orig=df_input[cols_orig].copy()
        df_orig.insert(len(df_orig.columns)-(1 if 'Reference' in df_orig.columns else 0),
                   'SUM_input',df_input[OXIDES].sum(axis=1).round(3))
        write_sheet(ws_in,df_orig,hdr_color='1B5E20')
        ws_rc=wb_chem.create_sheet('Recalculated_Chemistry')
        cols_rec=['Sample']+OXIDES+['SUM','Fe_treatment']
        if 'Reference' in rows_recalc[0]: cols_rec.append('Reference')
        write_sheet(ws_rc,pd.DataFrame(rows_recalc)[cols_rec],hdr_color='1F4E79')

        buf_visc=io.BytesIO(); wb.save(buf_visc); buf_visc.seek(0)
        buf_chem=io.BytesIO(); wb_chem.save(buf_chem); buf_chem.seek(0)
        buf_plot=io.BytesIO(); fig.savefig(buf_plot,format='png',dpi=200); buf_plot.seek(0)

        st.subheader("📥 Download results")
        c1,c2,c3=st.columns(3)
        with c1: st.download_button("⬇️ Viscosity Excel",data=buf_visc,
            file_name="output_viscosity.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c2: st.download_button("⬇️ Chemistry check Excel",data=buf_chem,
            file_name="chemistry_check.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        with c3: st.download_button("⬇️ Plot (PNG)",data=buf_plot,
            file_name="viscosity_plot.png",mime="image/png")

        if skipped:
            st.warning(f"⚠️ {len(skipped)} samples could not be processed:")
            st.dataframe(pd.DataFrame(skipped))
        else:
            st.success(f"✅ All {len(df_input)} samples processed successfully!")

    # ==============================================================================
    # MODE 2 — ANHYDROUS AND HYDROUS MODELLING
    # ==============================================================================
elif mode == "💧 Anhydrous and Hydrous Modelling":

    st.title("💧 Anhydrous and Hydrous Modelling")
    st.markdown("""
Model how **Tg, fragility index m and viscosity** evolve as a function of H₂O content.  
Melt viscosity is calculated using the Artificial Neural Network (ANN) of
[Langhammer et al. (2022)](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2022GC010673).  
The dependence of Tg and m on water content follows the framework of
[Langhammer et al. (2021)](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918)
(Eq. 9-10, 12).  
Upload a CSV with **one anhydrous composition** (H₂O = 0).
    """)

    with st.expander("📖 How to use this mode — read first", expanded=False):
        st.markdown("""
### What this mode does
This mode takes a single anhydrous melt composition and models how viscosity changes
as a function of dissolved water content, following the physically-motivated approach
of Langhammer et al. (2021).

### Step-by-step
1. **Upload a CSV** with one row containing your anhydrous composition in wt% oxides (H₂O = 0).
2. **Set H₂O contents** — the tool will recalculate the ANN viscosity at each water content
   you specify (e.g. 0, 0.5, 1, 2, 3, 4, 5 wt%). Always include 0.
3. **Click Run** — the ANN calculates Tg and m at each water content. Then:
   - **Tg(H₂O)** is fitted using Eq. 9-10 (Schneider/Gordon-Taylor model, three free parameters b, c, d)
   - **m(H₂O)** is modelled in three ways (see below)
   - **Viscosity curves** are generated for each m model
4. **Download** the Excel file with all results and the plot as PNG.

### Compositional limits
The ANN was trained on compositions with SiO₂ ranging from ~43 to 79 wt% and total
alkalis from 0 to 17 wt%. Results outside this range may be unreliable.
Always verify that your composition falls within the training domain of the model.

### What the diamond markers mean
The ♦ diamond on each viscosity curve marks the **glass transition temperature Tg**,
defined as the temperature at which log₁₀(η) = 12 Pa·s.
    """)

    # ── Input method selector ─────────────────────────────────────────────────
    input_method = st.radio("**How do you want to provide the composition?**",
                        ["📂 Upload CSV file", "⌨️ Type composition manually"],
                        horizontal=True, key="hyd_input_method")

    df_h = None

    if input_method == "📂 Upload CSV file":
        st.download_button("📄 Download CSV template", data=TEMPLATE_CSV,
                       file_name="Example_compositions.csv", mime="text/csv")
        uploaded_h = st.file_uploader("Upload anhydrous composition CSV", type=["csv"],
                                   key="hyd_upload")
        if uploaded_h is None:
            _default_csv_h = pathlib.Path(BASE_DIR) / "Example_compositions.csv"
            if _default_csv_h.exists():
                df_h = pd.read_csv(_default_csv_h).dropna(how='all').reset_index(drop=True)
            else:
                import io as _sio
                df_h = pd.read_csv(_sio.StringIO(DEFAULT_EXAMPLE_CSV)).dropna(how='all').reset_index(drop=True)
            st.info("📋 Using bundled **Example_compositions.csv** — upload your own CSV file above to replace it.")
        else:
            df_h = pd.read_csv(uploaded_h).dropna(how='all').reset_index(drop=True)
        for ox in OXIDES:
            if ox not in df_h.columns: df_h[ox] = 0.0
        df_h[OXIDES] = df_h[OXIDES].fillna(0.0)
        if len(df_h) == 0:
            st.error("No valid rows found in the CSV.")
            st.stop()
        if len(df_h) > 1:
            # ── TAS diagram for composition selection ─────────────────────────
            st.markdown("**Select composition — click on a point in the TAS or use the dropdown:**")

            import warnings; warnings.filterwarnings('ignore')
            from pyrolite.util.classification import TAS as TASclf
            import matplotlib.patheffects as _pe

            # Iron redistribution + normalize for TAS plotting
            df_h_norm = df_h.copy()
            for _idx, _row in df_h_norm.iterrows():
                _feo = _row['FeO']; _fe2o3 = _row['Fe2O3']
                if _feo > 0 and _fe2o3 == 0:
                    df_h_norm.at[_idx,'FeO'] = _feo/2; df_h_norm.at[_idx,'Fe2O3'] = _feo*1.11/2
                elif _fe2o3 > 0 and _feo == 0:
                    df_h_norm.at[_idx,'Fe2O3'] = _fe2o3/2; df_h_norm.at[_idx,'FeO'] = _fe2o3/1.11/2
                _s = df_h_norm.loc[_idx, OXIDES].sum()
                if _s > 0: df_h_norm.loc[_idx, OXIDES] = df_h_norm.loc[_idx, OXIDES] / _s * 100.0

            # Classify and get current selection
            _all_samples = df_h_norm['Sample'].tolist()
            _current_sel = st.session_state.get('hyd_sample_select', _all_samples[0])
            if _current_sel not in _all_samples:
                _current_sel = _all_samples[0]

            # ── TAS: matplotlib (beautiful) + Plotly points-only for clicking ─
            # ── Plotly TAS with exact pyrolite coordinates + clickable points ──
            import plotly.graph_objects as go
            from streamlit_plotly_events import plotly_events

            _sio2_vals = [float(df_h_norm.loc[df_h_norm['Sample']==s,'SiO2'].values[0])
                         for s in _all_samples]
            _tas_vals  = [(float(df_h_norm.loc[df_h_norm['Sample']==s,'Na2O'].values[0]) +
                          float(df_h_norm.loc[df_h_norm['Sample']==s,'K2O'].values[0]))
                         for s in _all_samples]

            # Extract exact polygon vertices + label positions from pyrolite at runtime
            import warnings as _w; _w.filterwarnings('ignore')
            _fig_ext, _ax_ext = plt.subplots()
            _tas_ext = TASclf()
            _tas_ext.add_to_axes(ax=_ax_ext, add_labels=True, which_labels='volcanic',
                                 facecolor='#F0F0F0', edgecolor='#888888')
            _pyro_polys  = [p.get_path().vertices.tolist() for p in _ax_ext.patches]
            _pyro_labels = [(t.get_text().replace("\n"," "), t.get_position())
                            for t in _ax_ext.texts]
            plt.close(_fig_ext)

            # Build Plotly figure
            _fig_p = go.Figure()

            # Polygons — exact pyrolite boundaries, styled like matplotlib
            for _verts in _pyro_polys:
                _xs = [v[0] for v in _verts]
                _ys = [v[1] for v in _verts]
                _fig_p.add_trace(go.Scatter(
                    x=_xs, y=_ys, fill='toself',
                    fillcolor='rgba(255,255,255,0.0)',
                    line=dict(color='#999999', width=1),
                    mode='lines', showlegend=False, hoverinfo='skip'
                ))

            # Field labels
            for _lbl, (_lx, _ly) in _pyro_labels:
                _fig_p.add_annotation(
                    x=_lx, y=_ly, text=_lbl, showarrow=False,
                    font=dict(size=8.5, color='#505050', family='Arial'),
                    bgcolor='rgba(255,255,255,0.0)', xanchor='center', yanchor='middle'
                )

            # Sample points — one trace per sample for reliable curveNumber
            _N_POLY = len(_pyro_polys)
            for _si, _sn in enumerate(_all_samples):
                _is_sel = (_sn == _current_sel)
                _fig_p.add_trace(go.Scatter(
                    x=[_sio2_vals[_si]], y=[_tas_vals[_si]],
                    mode='markers+text',
                    marker=dict(
                        color='tomato' if _is_sel else '#4C8CB5',
                        size=16 if _is_sel else 10,
                        line=dict(color='black', width=2 if _is_sel else 0.8),
                    ),
                    text=[_sn], textposition='top right',
                    textfont=dict(size=9, color='#111111', family='Arial'),
                    name=_sn, showlegend=False,
                    hovertemplate=(f'<b>{_sn}</b><br>SiO₂: {_sio2_vals[_si]:.1f} wt%'
                                   f'<br>Na₂O+K₂O: {_tas_vals[_si]:.1f} wt%<extra></extra>'),
                ))

            _fig_p.update_layout(
                xaxis=dict(title='SiO₂ (wt%)', range=[35, 82],
                           showgrid=True, gridcolor='rgba(200,200,200,0.5)',
                           gridwidth=1, zeroline=False, linecolor='#888',
                           ticks='outside', tickfont=dict(size=11)),
                yaxis=dict(title='Na₂O + K₂O (wt%)', range=[0, 16],
                           showgrid=True, gridcolor='rgba(200,200,200,0.5)',
                           gridwidth=1, zeroline=False, linecolor='#888',
                           ticks='outside', tickfont=dict(size=11)),
                title=dict(text=(f'TAS  ·  🔴 <b>{_current_sel}</b>'
                                 f'  · click a symbol to change'),
                           font=dict(size=12, family='Arial'), x=0.5),
                height=530, margin=dict(l=60, r=20, t=45, b=55),
                plot_bgcolor='white', paper_bgcolor='white',
                dragmode=False, font=dict(family='Arial')
            )

            _clicked = plotly_events(_fig_p, click_event=True,
                                     override_height=530, key='tas_click')
            if _clicked:
                _c = _clicked[0]
                _sample_idx = _c.get('curveNumber', -1) - _N_POLY
                if 0 <= _sample_idx < len(_all_samples):
                    _clicked_name = _all_samples[_sample_idx]
                    if _clicked_name != _current_sel:
                        st.session_state['hyd_sample_select'] = _clicked_name
                        _current_sel = _clicked_name

            # Dropdown — no key= so index= always reflects current selection
            selected_sample = st.selectbox(
                "Or select from dropdown:",
                options=_all_samples,
                index=_all_samples.index(_current_sel)
            )
            if selected_sample != _current_sel:
                st.session_state['hyd_sample_select'] = selected_sample
                _current_sel = selected_sample

            # Use current selection
            df_h = df_h[df_h["Sample"] == _current_sel].reset_index(drop=True)
            st.session_state['df_h_selected'] = df_h.to_dict('records')
            st.success(f"✅ Selected: **{_current_sel}**")
        else:
            df_h = df_h.iloc[[0]]
            st.session_state['df_h_selected'] = df_h.to_dict('records')

    else:  # Manual input
        st.markdown("**Enter your anhydrous composition in wt% (H₂O = 0):**")
        sample_name_manual = st.text_input("Sample name:", value="My_sample", key="hyd_sname")
        OXIDES_NO_H2O = [o for o in OXIDES if o != "H2O"]
        cols_manual = st.columns(4)
        defaults_m = {
            "SiO2": 48.05, "TiO2": 0.76, "Al2O3": 17.69, "FeO": 6.08,
            "MnO": 0.14, "MgO": 3.32, "CaO": 9.31, "Na2O": 3.45,
            "K2O": 7.55, "P2O5": 0.46, "Cr2O3": 0.0, "Fe2O3": 0.0,
        }
        manual_vals = {}
        for i, ox in enumerate(OXIDES_NO_H2O):
            with cols_manual[i % 4]:
                manual_vals[ox] = st.number_input(
                    ox, min_value=0.0, max_value=100.0,
                    value=float(defaults_m.get(ox, 0.0)),
                    step=0.01, format="%.3f",
                    key=f"hyd_manual_{ox}")
        manual_vals["H2O"] = 0.0
        total = sum(manual_vals[o] for o in OXIDES_NO_H2O)
        st.caption(f"Sum of oxides (anhydrous): **{total:.2f} wt%** — will be normalised to 100%")
        df_h = pd.DataFrame([{"Sample": sample_name_manual, **manual_vals}])
        st.session_state['df_h_selected'] = df_h.to_dict('records')
        st.success(f"✅ Composition ready: **{df_h['Sample'].values[0]}**")
        with st.expander("Preview composition"): st.dataframe(df_h)

    # H2O contents input
    st.subheader("⚙️ Settings")
    h2o_input = st.text_input(
        "H₂O contents to model (wt%, comma-separated):",
        value="0, 0.5, 1, 2, 3, 4, 5",
        help="Always include 0 as the anhydrous reference.")

    try:
        h2o_list = sorted(set([float(x.strip()) for x in h2o_input.split(',') if x.strip()]))
        if 0.0 not in h2o_list: h2o_list = [0.0] + h2o_list
        st.caption(f"Will calculate at: {h2o_list} wt% H₂O")
    except ValueError:
        st.error("Invalid H₂O input. Use numbers separated by commas.")
        st.stop()

    # Restore df_h from session state in case of button-click rerun
    if st.session_state.get('df_h_selected') and (df_h is None or len(df_h) == 0):
        df_h = pd.DataFrame(st.session_state['df_h_selected'])

    if st.button("▶️ Run hydrous modelling", type="primary"):

        model = load_model()
        row0 = df_h.iloc[0]
        sname = row0['Sample']
        wt_orig = np.array([row0[o] for o in OXIDES], dtype=float)
        wt_orig[12] = 0.0  # ensure anhydrous
        wt_dry, _ = redistribute_iron(wt_orig)
        wt_dry = normalize_to_100(wt_dry)
        st.session_state['wt_dry'] = wt_dry

        results = []
        progress = st.progress(0, text="Calculating ANN points...")

        for i, h2o in enumerate(h2o_list):
            try:
                Tg, m = get_Tg_m(wt_dry, h2o, model)
                x_mol = wt_to_mol_h2o(h2o, wt_dry)
                results.append({'h2o_wt': h2o, 'h2o_mol': x_mol, 'Tg': Tg, 'm': m})
            except Exception as e:
                st.warning(f"H₂O = {h2o} wt% failed: {e}")
            progress.progress((i+1)/len(h2o_list), text=f"Calculating H₂O = {h2o} wt%...")
        progress.empty()

        if len(results) < 2:
            st.error("Not enough points calculated. Check your composition.")
            st.stop()

        Tg_d = results[0]['Tg']
        m_d  = results[0]['m']
        x_mol_arr = np.array([r['h2o_mol'] for r in results])
        Tg_arr    = np.array([r['Tg']     for r in results])
        m_arr     = np.array([r['m']      for r in results])

        # Fit Tg(x_H2O) — Eq. 9-10
        with st.spinner("Fitting Tg(H₂O) with Eq. 9-10..."):
            (b_fit, c_fit, d_fit), tg_rmse = fit_tg(x_mol_arr, Tg_arr, Tg_d)

        # ── Physical filter for m ─────────────────────────────────────────────
        # Physical constraint: m(H2O) must be <= m_d (anhydrous value).
        # Adding water cannot increase fragility above the dry value.
        # Strategy:
        #   1. Exclude all points where m > m_d (unphysical, e.g. initial ANN overshoot)
        #   2. Among remaining points, keep only those forming a monotone
        #      decreasing sequence (first occurrence wins at each x)
        mono_mask = np.zeros(len(m_arr), dtype=bool)
        mono_mask[0] = True   # always keep x=0 anchor
        _m_running_min = m_d  # reference is m_d, not m_arr[0] which may be != m_d
        for _i in range(1, len(m_arr)):
            if m_arr[_i] <= _m_running_min:   # physically valid AND monotone
                mono_mask[_i] = True
                _m_running_min = m_arr[_i]
        non_mono_detected = not np.all(mono_mask)
        x_fit  = x_mol_arr[mono_mask]
        m_fit  = m_arr[mono_mask]
        # Safety: need at least 2 points for fit; if not, use all points <= m_d
        if mono_mask.sum() < 2:
            mono_mask = m_arr <= m_d
            mono_mask[0] = True
            x_fit = x_mol_arr[mono_mask]
            m_fit = m_arr[mono_mask]

        # ── Fit m with ANCHORED polynomial (passes through m_d at x=0) ───────
        from scipy.optimize import curve_fit as _cf
        dm_fit = m_fit - m_d

        def _anch_lin(x, alpha):
            return alpha * x
        def _anch_quad(x, alpha, beta):
            return alpha*x + beta*x**2

        # ── Exponential saturation fit: m(x) = m_inf + (m_d - m_inf)*exp(-k*x)
        # - Always anchored at m_d at x=0
        # - Always monotone decreasing if m_inf < m_d and k > 0
        # - Captures the steep initial drop + flattening at high H2O

        def _exp_sat(x, m_inf, k):
            return m_inf + (m_d - m_inf) * np.exp(-k * x)

        best_rmse = np.inf
        best_params = None
        for m_inf_0 in [m_d * 0.5, m_d * 0.6, m_d * 0.7, m_d * 0.8, 20.0, 25.0]:
            for k_0 in [0.1, 0.2, 0.3, 0.5, 1.0]:
                try:
                    popt, _ = _cf(_exp_sat, x_fit, m_fit,
                                  p0=[m_inf_0, k_0],
                                  bounds=([0, 1e-4], [m_d - 0.1, 10.0]),
                                  maxfev=2000)
                    r = np.sqrt(np.mean((_exp_sat(x_fit, *popt) - m_fit)**2))
                    if r < best_rmse:
                        best_rmse = r
                        best_params = popt
                except: pass

        if best_params is not None:
            _m_inf = float(best_params[0])
            _k     = float(best_params[1])
        else:
            # Fallback: linear anchored
            _m_inf = float(m_d * 0.7)
            _k     = 0.2

        m_poly_anchored = lambda x, mi=_m_inf, k=_k, md=m_d: mi + (md - mi)*np.exp(-k*x)
        poly_deg   = 1   # label only
        _alpha     = _k
        _beta      = _m_inf
        poly_rmse  = best_rmse
        poly_label = 'm_inf + (m_d − m_inf)·exp(−k·x)  [m_inf={:.2f}, k={:.4f}]'.format(_m_inf, _k)
        m_poly = m_poly_anchored

        # Save to session state and warn if non-monotone behaviour detected
        st.session_state['non_mono'] = non_mono_detected

        # Smooth curves
        x_smooth   = np.linspace(0, max(x_mol_arr)*1.05, 200)
        Tg_smooth  = np.array([tg_model(x, b_fit, c_fit, d_fit, Tg_d) for x in x_smooth])
        m_eq12_sm  = np.array([m_from_tg(Tg, Tg_d, m_d) for Tg in Tg_smooth])
        m_poly_sm  = np.array([m_poly(x) for x in x_smooth])

        # ── Figure: 5 panels ─────────────────────────────────────────────────
        colors_visc = cm.plasma(np.linspace(0.1, 0.9, len(results)))
        fig2, axes = plt.subplots(1, 5, figsize=(26, 5))
        fig2.suptitle(f"Anhydrous & Hydrous Modelling — {sname}", fontsize=13, fontweight='bold')

        # Panel 1: Tg vs H2O
        ax1 = axes[0]
        ax1.scatter(x_mol_arr, Tg_arr-273.15, color='steelblue', s=60, zorder=5, label='ANN')
        ax1.plot(x_smooth, Tg_smooth-273.15, 'steelblue', linewidth=2,
                 label='Eq.9-10 fit\nb={:.3f}\nc={:.3f}\nd={:.3f}'.format(b_fit,c_fit,d_fit))
        ax1.set_xlabel('H$_2$O (mol%)', fontsize=11)
        ax1.set_ylabel('Tg (°C)', fontsize=11)
        ax1.set_title('Glass transition temperature', fontsize=11)
        ax1.legend(fontsize=7, loc='lower left'); ax1.grid(True, linestyle='--', alpha=0.4)

        # Panel 2: m vs H2O
        ax2 = axes[1]
        # Same physical filter
        _mm = np.zeros(len(m_arr), dtype=bool)
        _mm[0] = True
        _mn = m_d
        for _ii in range(1, len(m_arr)):
            if m_arr[_ii] <= _mn: _mm[_ii] = True; _mn = m_arr[_ii]
        ax2.scatter(x_mol_arr[_mm], m_arr[_mm], color='tomato', s=60, zorder=5,
                    label='ANN (used for fit)')
        if not np.all(_mm):
            ax2.scatter(x_mol_arr[~_mm], m_arr[~_mm], color='lightgray', s=60,
                        marker='x', linewidths=2, zorder=5, label='ANN (excluded)')
        ax2.plot(x_smooth, m_eq12_sm, 'tomato', linewidth=2, linestyle='--', label='Eq. 12, Langhammer et al. (2021)')
        ax2.plot(x_smooth, m_poly_sm, 'darkorange', linewidth=2, label='Exp. saturation (ANN)')
        ax2.axhline(m_d, color='steelblue', linewidth=2, linestyle=':', label='m constant = {:.2f}'.format(m_d))
        ax2.set_xlabel('H$_2$O (mol%)', fontsize=11)
        ax2.set_ylabel('Fragility index m', fontsize=11)
        ax2.set_title('Fragility index — three models', fontsize=11)
        ax2.legend(fontsize=7, loc='lower left'); ax2.grid(True, linestyle='--', alpha=0.4)

        # Helper
        def visc_panel(ax, m_func, title):
            for i, r in enumerate(results):
                Tg_f = tg_model(r['h2o_mol'], b_fit, c_fit, d_fit, Tg_d)
                m_v  = m_func(r['h2o_mol'])
                try: T_max = brentq(myega_eq, Tg_f, 5000.0, args=(Tg_f, m_v))
                except: T_max = 3000.0
                T_arr2 = np.arange(Tg_f, T_max+50, 25)
                ax.plot(T_arr2-273.15, myega_eq(T_arr2, Tg_f, m_v),
                        color=colors_visc[i], linewidth=2,
                        label='{:.1f} wt%'.format(r['h2o_wt']))
            ax.set_xlabel('Temperature (°C)', fontsize=11)
            ax.set_ylabel('log$_{10}$(η / Pa·s)', fontsize=11)
            ax.set_title(title, fontsize=11)
            ax.legend(fontsize=7, loc='upper right')
            ax.grid(True, linestyle='--', alpha=0.4)

        visc_panel(axes[2], lambda x: m_d,
                   'm constant = {:.2f}'.format(m_d))
        visc_panel(axes[3],
                   lambda x: m_from_tg(tg_model(x, b_fit, c_fit, d_fit, Tg_d), Tg_d, m_d),
                   'm from Eq. 12, Langhammer et al. (2021)')
        visc_panel(axes[4],
                   lambda x: float(m_poly(x)),
                   'Exp. saturation (ANN fit)')

        plt.tight_layout()

        # ── Excel ─────────────────────────────────────────────────────────────
        wb2 = Workbook()

        # Sheet 1: Parameters
        ws_p = wb2.active; ws_p.title = 'Parameters'
        thin2=Side(style='thin',color='AAAAAA')
        brd2=Border(left=thin2,right=thin2,top=thin2,bottom=thin2)
        ctr2=Alignment(horizontal='center',vertical='center')
        def hdr2(cell, color='1B5E20'):
            cell.font=Font(name='Arial',bold=True,color='FFFFFF',size=10)
            cell.fill=PatternFill('solid',start_color=color)
            cell.alignment=Alignment(horizontal='center',vertical='center',wrap_text=True)
            cell.border=brd2
        def dat2(cell, alt=False, col=None):
            if col == 3:  # Description column: left-align + wrap
                cell.alignment=Alignment(horizontal='left',vertical='center',wrap_text=True)
            elif col == 1:  # Parameter column: left-align
                cell.alignment=Alignment(horizontal='left',vertical='center')
            else:
                cell.alignment=ctr2
            cell.border=brd2
            if alt: cell.fill=PatternFill('solid',start_color='D6E4F0')
            if isinstance(cell.value,float): cell.number_format='0.000000'

        param_data = [
            # ── Sample info ──────────────────────────────────────────────────────
            ('Sample',          sname,                   'Sample name'),
            # ── MYEGA parameters (Mauro et al. 2009) ────────────────────────────
            ('A',               A_FIXED,                 'Pre-exponential term A in MYEGA eq. [log Pa.s] — fixed at -2.9 (Langhammer et al. 2021)'),
            # ── Anhydrous Tg and m (ANN) ─────────────────────────────────────────
            ('Tg_dry (K)',      round(Tg_d,2),           'Anhydrous glass transition temperature [K] — ANN output'),
            ('Tg_dry (°C)',     round(Tg_d-273.15,2),    'Anhydrous glass transition temperature [°C] — ANN output'),
            ('m_dry',           round(m_d,3),            'Anhydrous fragility index m — ANN output  |  m = (12-A) / log(Tg/T_onset)'),
            # ── Tg(H2O) fit — Gordon-Taylor / Schneider (Eqs. 9-10, Langhammer 2021) ──
            ('',                '',                      '--- Tg(H2O) fit: Eq.9-10 Langhammer et al. (2021) ---'),
            ('b  (Eq.10)',      round(b_fit,5),          'Fit parameter b in Gordon-Taylor equation for Tg(H2O)'),
            ('c  (Eq.10)',      round(c_fit,5),          'Fit parameter c (accounts for excess mixing)'),
            ('d  (Eq.10)',      round(d_fit,5),          'Fit parameter d (higher-order correction)'),
            ('Tg fit RMSE (K)', round(tg_rmse,3),       'Root-mean-square error of Tg(H2O) fit [K]'),
            # ── m(H2O) exponential saturation fit ───────────────────────────────
            ('',                '',                      '--- m(H2O) fit: exponential saturation model ---'),
            ('m_inf',           round(_m_inf, 4),        'Limiting fragility at high H2O content  |  m --> m_inf as H2O --> inf'),
            ('k  [mol%-1]',     round(_k, 6),            'Decay rate constant  |  controls how fast m drops from m_dry to m_inf'),
            ('m fit RMSE',      round(poly_rmse, 4) if poly_rmse < 999 else 'n/a',
                                                         'Root-mean-square error of m(H2O) exponential saturation fit'),
            ('m(H2O) equation', 'm(x) = m_inf + (m_dry - m_inf) * exp(-k*x)',
                                                         'x = H2O in mol%;  anchored at m_dry at x=0;  always monotone decreasing'),
        ]
        col_widths = {'A': 22, 'B': 30, 'C': 70}
        for c,h in enumerate(['Parameter','Value','Description'],1):
            cell=ws_p.cell(row=1,column=c,value=h)
            hdr2(cell,'1B5E20')
            ws_p.column_dimensions[get_column_letter(c)].width=col_widths[get_column_letter(c)]
        ws_p.row_dimensions[1].height=25
        for r,(p,v,d) in enumerate(param_data,2):
            ws_p.row_dimensions[r].height = 30
            for c,val in enumerate([p,v,d],1):
                cell=ws_p.cell(row=r,column=c,value=val)
                dat2(cell,alt=(r%2==0), col=c)

        # Sheet 2: Tg and m vs H2O
        ws_tgm = wb2.create_sheet('Tg_m_vs_H2O')
        h2o_headers=['H2O (wt%)','H2O (mol%)','Tg_ANN (K)','Tg_ANN (C)',
                     'Tg_fit (K)','Tg_fit (C)','m_ANN','m_Eq12','m_ExpSat','m_constant']
        for c,h in enumerate(h2o_headers,1):
            cell=ws_tgm.cell(row=1,column=c,value=h)
            hdr2(cell,'4A235A')
            ws_tgm.column_dimensions[get_column_letter(c)].width=max(len(h),8)+3
        ws_tgm.row_dimensions[1].height=25
        for r,res_r in enumerate(results,2):
            Tg_f   = tg_model(res_r['h2o_mol'], b_fit, c_fit, d_fit, Tg_d)
            m_eq12 = m_from_tg(Tg_f, Tg_d, m_d)
            m_p    = float(m_poly(res_r['h2o_mol']))
            vals   = [round(res_r['h2o_wt'],2), round(res_r['h2o_mol'],3),
                      round(res_r['Tg'],2),      round(res_r['Tg']-273.15,2),
                      round(Tg_f,2),             round(Tg_f-273.15,2),
                      round(res_r['m'],3),        round(m_eq12,3),
                      round(m_p,3),              round(m_d,3)]
            for c,val in enumerate(vals,1):
                cell=ws_tgm.cell(row=r,column=c,value=val)
                dat2(cell,alt=(r%2==0))
        ws_tgm.freeze_panes='A2'

        # Sheet 3: Chemistry check
        ws_chem = wb2.create_sheet('Chemistry_Check')
        chem_headers = (['Field'] + OXIDES +
                        ['SUM_input', 'SUM_normalised', 'Fe_treatment'] +
                        [o+'_mol%' for o in OXIDES])
        for c, h in enumerate(chem_headers, 1):
            cell = ws_chem.cell(row=1, column=c, value=h)
            hdr2(cell, '1B5E20')
            ws_chem.column_dimensions[get_column_letter(c)].width = max(len(h), 6) + 2
        ws_chem.row_dimensions[1].height = 28

        # Row 1: Input composition (as entered by user)
        wt_input_row = np.array([row0[o] for o in OXIDES], dtype=float)
        wt_input_row[12] = 0.0  # anhydrous
        sum_input = wt_input_row.sum()

        # Row 2: After iron redistribution
        wt_fe_row, fe_flag_row = redistribute_iron(wt_input_row)

        # Row 3: Normalised to 100% (used for ANN)
        wt_norm_row = normalize_to_100(wt_fe_row)
        sum_norm = wt_norm_row.sum()

        # Mol% conversion
        _, _, mol_per_row = mol_conv(wt_norm_row)

        rows_chem = [
            ['Input (wt%, as entered)'] +
            [round(float(v), 4) for v in wt_input_row] +
            [round(sum_input, 4), '—', '—'] +
            ['—'] * len(OXIDES),

            ['After Fe redistribution (wt%)'] +
            [round(float(v), 4) for v in wt_fe_row] +
            [round(wt_fe_row.sum(), 4), '—', fe_flag_row] +
            ['—'] * len(OXIDES),

            ['Normalised to 100% — used for ANN (wt%)'] +
            [round(float(v), 4) for v in wt_norm_row] +
            [round(sum_input, 4), round(sum_norm, 4), fe_flag_row] +
            [round(float(v), 4) for v in mol_per_row],
        ]

        color_map = ['FFFFFF', 'FFF9C4', 'E3F2FD']
        for r, row_data in enumerate(rows_chem, 2):
            fill_c = PatternFill('solid', start_color=color_map[r-2])
            for c, val in enumerate(row_data, 1):
                cell = ws_chem.cell(row=r, column=c, value=val)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                thin3 = Side(style='thin', color='AAAAAA')
                cell.border = Border(left=thin3, right=thin3, top=thin3, bottom=thin3)
                cell.fill = fill_c
                if isinstance(val, float): cell.number_format = '0.0000'

        # Add a legend below
        legend_row = 6
        ws_chem.cell(row=legend_row, column=1,
            value="Row colours: White = raw input | Yellow = after Fe redistribution | Blue = normalised (used for ANN + mol% conversion)"
        ).font = Font(name='Arial', italic=True, size=8, color='555555')

        ws_chem.freeze_panes = 'B2'

        # Sheets 4-6: viscosity curves
        tg_func = lambda x: tg_model(x, b_fit, c_fit, d_fit, Tg_d)
        make_visc_sheet_hydrous(wb2,'Visc_m_constant',
            results, lambda x: m_d, tg_func, '1F4E79')
        make_visc_sheet_hydrous(wb2,'Visc_m_Eq12',
            results, lambda x: m_from_tg(tg_func(x), Tg_d, m_d), tg_func, '7B1FA2')
        make_visc_sheet_hydrous(wb2,'Visc_m_ExpSat',
            results, lambda x: float(m_poly(x)), tg_func, 'BF360C')

        buf_excel = io.BytesIO(); wb2.save(buf_excel); buf_excel.seek(0)
        buf_fig2  = io.BytesIO(); fig2.savefig(buf_fig2,format='png',dpi=200,bbox_inches='tight')
        buf_fig2.seek(0)

        st.session_state['hyd_done']     = True
        st.session_state['hyd_results']  = results
        st.session_state['hyd_fig']      = fig2
        st.session_state['hyd_buf_excel']= buf_excel
        st.session_state['hyd_buf_fig']  = buf_fig2
        # Store poly params to reconstruct m_poly after page reload
        _poly_params = {'m_inf': _m_inf, 'k': _k}
        st.session_state['hyd_meta']     = {
            'sname':sname, 'Tg_d':Tg_d, 'm_d':m_d,
            'b':b_fit, 'c':c_fit, 'd':d_fit,
            'tg_rmse':tg_rmse, 'poly_deg':poly_deg,
            'poly_params': _poly_params,
            'poly_label': poly_label,
        }

    # ── Show results ──────────────────────────────────────────────────────────
    if st.session_state['hyd_done']:
        meta = st.session_state['hyd_meta']

        st.subheader("📈 Results")
        st.markdown(f"""
**Tg_dry** = {meta['Tg_d']-273.15:.1f} °C &nbsp;|&nbsp;
**m_dry** = {meta['m_d']:.2f} &nbsp;|&nbsp;
**Tg fit:** b={meta['b']:.4f}, c={meta['c']:.4f}, d={meta['d']:.4f} &nbsp;|&nbsp;
**RMSE** = {meta['tg_rmse']:.2f} K
        """)
        if st.session_state.get('non_mono'):
            st.warning("⚠️ **Non-monotone m (ANN)**: the ANN produces m values that are not monotonically decreasing with H₂O content. Only monotone decreasing points were used for the exponential saturation fit. See Langhammer et al. (2021).")
        st.pyplot(st.session_state['hyd_fig'])

        # ── Individual panel downloads ─────────────────────────────────────────
        st.subheader("📥 Download individual panels")
        results_ss   = st.session_state['hyd_results']
        meta_ss      = st.session_state['hyd_meta']
        Tg_d_ss      = meta_ss['Tg_d']
        m_d_ss       = meta_ss['m_d']
        b_ss         = meta_ss['b']
        c_ss         = meta_ss['c']
        d_ss         = meta_ss['d']
        poly_deg_ss  = meta_ss['poly_deg']
        _pp    = meta_ss['poly_params']
        _mi_ss = _pp['m_inf']
        _k_ss  = _pp['k']
        _md_ss = meta_ss['m_d']
        m_poly_ss = lambda x, mi=_mi_ss, k=_k_ss, md=_md_ss: mi + (md - mi)*np.exp(-k*x)
        sname_ss     = meta_ss['sname']
        x_mol_arr_ss = np.array([r['h2o_mol'] for r in results_ss])
        Tg_arr_ss    = np.array([r['Tg']      for r in results_ss])
        m_arr_ss     = np.array([r['m']       for r in results_ss])
        x_smooth_ss  = np.linspace(0, max(x_mol_arr_ss)*1.05, 200)
        Tg_smooth_ss = np.array([tg_model(x, b_ss, c_ss, d_ss, Tg_d_ss) for x in x_smooth_ss])
        m_eq12_sm_ss = np.array([m_from_tg(Tg, Tg_d_ss, m_d_ss) for Tg in Tg_smooth_ss])
        m_poly_sm_ss = np.array([m_poly_ss(x) for x in x_smooth_ss])
        colors_ss    = cm.plasma(np.linspace(0.1, 0.9, len(results_ss)))

        tg_func_ss   = lambda x: tg_model(x, b_ss, c_ss, d_ss, Tg_d_ss)
        m_eq12_func  = lambda x: m_from_tg(tg_func_ss(x), Tg_d_ss, m_d_ss)
        m_poly_func  = lambda x: float(m_poly_ss(x))
        m_const_func = lambda x: m_d_ss

        def make_single_fig(panel_func, title):
            fig_s, ax_s = plt.subplots(figsize=(7, 5))
            fig_s.suptitle(f"{sname_ss} — {title}", fontsize=11, fontweight='bold')
            panel_func(ax_s)
            plt.tight_layout()
            buf = io.BytesIO()
            fig_s.savefig(buf, format='png', dpi=200, bbox_inches='tight')
            buf.seek(0)
            plt.close(fig_s)
            return buf

        def panel_tg(ax):
            ax.scatter(x_mol_arr_ss, Tg_arr_ss-273.15, color='steelblue', s=60, zorder=5, label='ANN')
            ax.plot(x_smooth_ss, Tg_smooth_ss-273.15, 'steelblue', linewidth=2,
                    label='Eq.9-10 fit  b={:.3f}, c={:.3f}, d={:.3f}'.format(b_ss,c_ss,d_ss))
            ax.set_xlabel('H₂O (mol%)'); ax.set_ylabel('Tg (°C)')
            ax.set_title('Glass transition temperature'); ax.legend(fontsize=8, loc='lower left'); ax.grid(True,linestyle='--',alpha=0.4)

        def panel_m(ax):
            # Same physical filter as in computation
            _mono_mask_ss = np.zeros(len(m_arr_ss), dtype=bool)
            _mono_mask_ss[0] = True
            _m_run_ss = meta_ss['m_d']
            for _ii in range(1, len(m_arr_ss)):
                if m_arr_ss[_ii] <= _m_run_ss:
                    _mono_mask_ss[_ii] = True; _m_run_ss = m_arr_ss[_ii]
            ax.scatter(x_mol_arr_ss[_mono_mask_ss], m_arr_ss[_mono_mask_ss],
                       color='tomato', s=60, zorder=5, label='ANN (used for fit)')
            if not np.all(_mono_mask_ss):
                ax.scatter(x_mol_arr_ss[~_mono_mask_ss], m_arr_ss[~_mono_mask_ss],
                           color='lightgray', s=60, marker='x', linewidths=2,
                           zorder=5, label='ANN (excluded)')
            ax.plot(x_smooth_ss, m_eq12_sm_ss, 'tomato', linewidth=2, linestyle='--', label='Eq. 12, Langhammer et al. (2021)')
            ax.plot(x_smooth_ss, m_poly_sm_ss, 'darkorange', linewidth=2, label='Exp. saturation (ANN)')
            ax.axhline(m_d_ss, color='steelblue', linewidth=2, linestyle=':', label='m constant = {:.2f}'.format(m_d_ss))
            ax.set_xlabel('H₂O (mol%)'); ax.set_ylabel('Fragility index m')
            ax.set_title('Fragility index'); ax.legend(fontsize=8, loc='lower left'); ax.grid(True,linestyle='--',alpha=0.4)

        def panel_visc(ax, m_func, title):
            for i, r in enumerate(results_ss):
                Tg_f = tg_func_ss(r['h2o_mol'])
                m_v  = m_func(r['h2o_mol'])
                try: T_max = brentq(myega_eq, Tg_f, 5000.0, args=(Tg_f, m_v))
                except: T_max = 3000.0
                T_arr2 = np.arange(Tg_f, T_max+50, 25)
                ax.plot(T_arr2-273.15, myega_eq(T_arr2,Tg_f,m_v), color=colors_ss[i],
                        linewidth=2, label='{:.1f} wt%'.format(r['h2o_wt']))
            ax.set_xlabel('Temperature (°C)'); ax.set_ylabel('log₁₀(η / Pa·s)')
            ax.set_title(title); ax.legend(fontsize=7,loc='upper right'); ax.grid(True,linestyle='--',alpha=0.4)

        col_dl = st.columns(5)
        panels = [
            ("Tg_vs_H2O",      "Tg vs H₂O",             panel_tg,  None),
            ("m_vs_H2O",       "m vs H₂O",               panel_m,   None),
            ("Visc_m_constant","Visc — m constant",       None,      m_const_func),
            ("Visc_m_Eq12",    "Visc — m Eq.12, Langhammer et al. (2021)",          None,      m_eq12_func),
            ("Visc_m_ExpSat",    "Visc — m Exp.Sat.",           None,      m_poly_func),
        ]
        for col, (fname, label, pf, mf) in zip(col_dl, panels):
            if pf is not None:
                buf_s = make_single_fig(pf, label)
            else:
                buf_s = make_single_fig(lambda ax, mf=mf, label=label: panel_visc(ax, mf, label), label)
            with col:
                st.download_button(f"⬇️ {label}", data=buf_s,
                    file_name=f"{sname_ss}_{fname}.png", mime="image/png",
                    key=f"dl_{fname}")

        # ── Viscosity vs H2O at fixed T ───────────────────────────────────────
        st.divider()
        st.subheader("📥 Download all results")
        c1_main, c2_main = st.columns(2)
        with c1_main:
            st.download_button("⬇️ Download Excel (all models)",
                data=st.session_state['hyd_buf_excel'],
                file_name="hydrous_visc_{}.xlsx".format(meta['sname']),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_excel_main")
        with c2_main:
            st.download_button("⬇️ Download combined plot (PNG)",
                data=st.session_state['hyd_buf_fig'],
                file_name="hydrous_visc_{}.png".format(meta['sname']),
                mime="image/png", key="dl_fig_main")

        st.divider()
        st.subheader("🌡️ Viscosity vs H₂O at a fixed temperature")
        st.markdown("""
Do you want to simulate how viscosity changes with decreasing water content
at a fixed temperature? This generates three plots — one per fragility model —
showing **log₁₀(η)** as a function of **H₂O content (wt%)** at the temperature you choose.
        """)

        do_visc_vs_h2o = st.checkbox("Yes, calculate viscosity vs H₂O at fixed T")

        if do_visc_vs_h2o:
            col_t1, col_t2, col_t3 = st.columns(3)
            with col_t1:
                T_fixed_C = st.number_input("Temperature (°C):", min_value=200.0,
                                            max_value=2000.0, value=1200.0, step=50.0)
            with col_t2:
                h2o_min = st.number_input("H₂O min (wt%):", min_value=0.0,
                                          max_value=10.0, value=0.0, step=0.1)
            with col_t3:
                h2o_max = st.number_input("H₂O max (wt%):", min_value=0.1,
                                          max_value=15.0, value=5.0, step=0.1)

            n_points = st.slider("Number of H₂O points:", min_value=5, max_value=50, value=20)

            if st.button("▶️ Calculate viscosity vs H₂O", key="btn_visc_h2o"):
                T_fixed_K = T_fixed_C + 273.15
                h2o_range = np.linspace(h2o_min, h2o_max, n_points)

                # Convert wt% to mol% using fitted Tg curve range
                wt_dry = st.session_state.get('wt_dry')
                h2o_mol_range = np.array([wt_to_mol_h2o(h, wt_dry) for h in h2o_range])

                # Calculate viscosity for each m model at each H2O content
                visc_const, visc_eq12, visc_poly = [], [], []
                for h2o_mol in h2o_mol_range:
                    Tg_f = tg_func_ss(h2o_mol)
                    for visc_list, m_func in [
                        (visc_const, m_const_func),
                        (visc_eq12,  m_eq12_func),
                        (visc_poly,  m_poly_func),
                    ]:
                        m_v = m_func(h2o_mol)
                        try:
                            v = float(myega_eq(T_fixed_K, Tg_f, m_v))
                            if not np.isfinite(v): v = np.nan
                        except: v = np.nan
                        visc_list.append(v)

                st.session_state['visc_h2o'] = {
                    'h2o_wt': h2o_range,
                    'h2o_mol': h2o_mol_range,
                    'visc_const': visc_const,
                    'visc_eq12':  visc_eq12,
                    'visc_poly':  visc_poly,
                    'T_fixed_C':  T_fixed_C,
                    'sname':      sname_ss,
                    'poly_deg':   poly_deg_ss,
                }

        if st.session_state.get('visc_h2o'):
            vh = st.session_state['visc_h2o']
            h2o_wt   = vh['h2o_wt']
            h2o_mol  = vh['h2o_mol']
            T_label  = vh['T_fixed_C']
            sname_v  = vh['sname']
            pd_v     = vh['poly_deg']

            model_specs = [
                ('m_constant', f'm = constant = {m_d_ss:.2f}',     vh['visc_const'], 'steelblue'),
                ('m_Eq12',     'm Eq.12, Langhammer et al. (2021)',                      vh['visc_eq12'],  'tomato'),
                (f'm_poly_deg{pd_v}', 'Exp. saturation (ANN)',   vh['visc_poly'],  'darkorange'),
            ]

            st.markdown("**Results at T = {:.0f} °C**".format(T_label))

            from scipy.interpolate import PchipInterpolator

            def _smooth_fit(x_arr, y_arr, n=300):
                """Monotone cubic spline (PCHIP) — passes exactly through all data points."""
                valid = np.isfinite(y_arr)
                if valid.sum() < 2: return x_arr[valid], y_arr[valid]
                interp = PchipInterpolator(x_arr[valid], y_arr[valid])
                x_sm = np.linspace(x_arr[valid][0], x_arr[valid][-1], n)
                return x_sm, interp(x_sm)

            cols_vh = st.columns(3)
            bufs_vh = []
            for col, (mkey, mlabel, visc_vals, mcolor) in zip(cols_vh, model_specs):
                varr = np.array(visc_vals, dtype=float)
                x_sm, y_sm = _smooth_fit(h2o_wt, varr)
                fig_vh, ax_vh = plt.subplots(figsize=(6, 5))
                ax_vh.plot(x_sm, y_sm, color=mcolor, linewidth=2.5)
                ax_vh.scatter(h2o_wt[np.isfinite(varr)], varr[np.isfinite(varr)],
                              color=mcolor, s=40, zorder=5)
                ax_vh.set_xlabel("H$_2$O (wt%)", fontsize=12)
                ax_vh.set_ylabel(r'log$_{10}$($\eta$ / Pa$\cdot$s)', fontsize=12)
                ax_vh.set_title("{} - T = {:.0f} C - {}".format(sname_v, T_label, mlabel),
                                fontsize=10, fontweight="bold")
                ax_vh.grid(True, linestyle="--", alpha=0.5)
                plt.tight_layout()
                with col: st.pyplot(fig_vh)
                buf_vh_i = io.BytesIO()
                fig_vh.savefig(buf_vh_i, format="png", dpi=200, bbox_inches="tight")
                buf_vh_i.seek(0)
                bufs_vh.append((mkey, buf_vh_i))
                plt.close(fig_vh)

            # Combined figure
            fig_comb, axes_comb = plt.subplots(1, 3, figsize=(18, 5))
            fig_comb.suptitle("{} - Viscosity vs H2O at T = {:.0f} C".format(sname_v, T_label),
                              fontsize=13, fontweight="bold")
            for ax_c, (mkey, mlabel, visc_vals, mcolor) in zip(axes_comb, model_specs):
                varr = np.array(visc_vals, dtype=float)
                x_sm, y_sm = _smooth_fit(h2o_wt, varr)
                ax_c.plot(x_sm, y_sm, color=mcolor, linewidth=2.5)
                ax_c.scatter(h2o_wt[np.isfinite(varr)], varr[np.isfinite(varr)],
                             color=mcolor, s=40, zorder=5)
                ax_c.set_xlabel("H$_2$O (wt%)", fontsize=11)
                ax_c.set_ylabel(r'log$_{10}$($\eta$ / Pa$\cdot$s)', fontsize=11)
                ax_c.set_title(mlabel, fontsize=11)
                ax_c.grid(True, linestyle="--", alpha=0.5)
            plt.tight_layout()
            buf_comb = io.BytesIO()
            fig_comb.savefig(buf_comb, format="png", dpi=200, bbox_inches="tight")
            buf_comb.seek(0)
            plt.close(fig_comb)

            # Excel with raw data
            wb_vh = Workbook()
            ws_vh = wb_vh.active; ws_vh.title = "Visc_vs_H2O"
            vh_headers = ["H2O (wt%)", "H2O (mol%)",
                          "log10_visc_m_constant", "log10_visc_m_Eq12", "log10_visc_m_poly"]
            thin_vh = Side(style="thin", color="AAAAAA")
            brd_vh  = Border(left=thin_vh, right=thin_vh, top=thin_vh, bottom=thin_vh)
            ctr_vh  = Alignment(horizontal="center", vertical="center")
            for c, h in enumerate(vh_headers, 1):
                cell = ws_vh.cell(row=1, column=c, value=h)
                cell.font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
                cell.fill  = PatternFill("solid", start_color="1F4E79")
                cell.alignment = Alignment(horizontal="center",vertical="center",wrap_text=True)
                cell.border = brd_vh
                ws_vh.column_dimensions[get_column_letter(c)].width = max(len(h),8)+3
            ws_vh.row_dimensions[1].height = 28
            alt_vh = PatternFill("solid", start_color="D6E4F0")
            wht_vh = PatternFill("solid", start_color="FFFFFF")
            vc_arr  = vh["visc_const"]
            veq_arr = vh["visc_eq12"]
            vp_arr  = vh["visc_poly"]
            for r,(hw,hm,vc,veq,vp) in enumerate(
                    zip(h2o_wt, h2o_mol, vc_arr, veq_arr, vp_arr), 2):
                fill_vh = alt_vh if r%2==0 else wht_vh
                for c,val in enumerate([round(float(hw),3), round(float(hm),3),
                                        round(float(vc),4), round(float(veq),4),
                                        round(float(vp),4)], 1):
                    cell = ws_vh.cell(row=r, column=c, value=val)
                    cell.alignment=ctr_vh; cell.border=brd_vh
                    cell.fill=fill_vh; cell.number_format="0.000"
            ws_vh.freeze_panes = "A2"
            buf_xl_vh = io.BytesIO(); wb_vh.save(buf_xl_vh); buf_xl_vh.seek(0)

            # 4th plot: all three models together
            fig_all, ax_all = plt.subplots(figsize=(7, 5))
            fig_all.suptitle("{} - T = {:.0f} C - All models".format(sname_v, T_label),
                             fontsize=11, fontweight="bold")
            ls_list = ["-", "--", ":"]
            for (mkey, mlabel, visc_vals, mcolor), ls in zip(model_specs, ls_list):
                varr4 = np.array(visc_vals, dtype=float)
                x_sm4, y_sm4 = _smooth_fit(h2o_wt, varr4)
                ax_all.plot(x_sm4, y_sm4, color=mcolor, linewidth=2.5, linestyle=ls, label=mlabel)
                ax_all.scatter(h2o_wt[np.isfinite(varr4)], varr4[np.isfinite(varr4)],
                               color=mcolor, s=40, zorder=5)
            ax_all.set_xlabel("H$_2$O (wt%)", fontsize=12)
            ax_all.set_ylabel(r"log$_{10}$($\eta$ / Pa$\cdot$s)", fontsize=12)
            ax_all.legend(fontsize=9); ax_all.grid(True, linestyle="--", alpha=0.5)
            plt.tight_layout()
            st.pyplot(fig_all)
            buf_all = io.BytesIO()
            fig_all.savefig(buf_all, format="png", dpi=200, bbox_inches="tight")
            buf_all.seek(0); plt.close(fig_all)

            st.subheader("Download viscosity vs H2O")
            dl_cols = st.columns(5)
            with dl_cols[0]:
                st.download_button("Combined (PNG)", data=buf_comb,
                    file_name="{}_visc_vs_H2O_T{:.0f}C_combined.png".format(sname_v, T_label),
                    mime="image/png", key="dl_vh_combined")
            for i,(mkey,buf_vh_i) in enumerate(bufs_vh):
                with dl_cols[i+1]:
                    st.download_button("{}".format(model_specs[i][1][:18]),
                        data=buf_vh_i,
                        file_name="{}_visc_vs_H2O_T{:.0f}C_{}.png".format(sname_v,T_label,mkey),
                        mime="image/png", key="dl_vh_{}".format(mkey))
            with dl_cols[4]:
                st.download_button("Excel data",
                    data=buf_xl_vh,
                    file_name="{}_visc_vs_H2O_T{:.0f}C.xlsx".format(sname_v, T_label),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_vh_excel")
            st.download_button("All models combined (PNG)",
                data=buf_all,
                file_name="{}_visc_vs_H2O_T{:.0f}C_all_models.png".format(sname_v, T_label),
                mime="image/png", key="dl_vh_all_models")

# ==============================================================================
# MODE 3 — SPECIFIC COMPOSITION MODELS
# ==============================================================================
elif mode == "🌋 Specific Composition Models":

    st.title("🌋 Specific Composition Models")
    st.markdown("""
Viscosity models calibrated on specific compositions:

- **Stromboli basalt** — [Valdivia et al. (2023)](https://link.springer.com/article/10.1007/s00410-023-02024-w)
- **Peridotite** — [Di Genova et al. (2023)](https://www.sciencedirect.com/science/article/pii/S0009254123001407)
- **Anhydrous andesite** — [Valdivia et al. (2025)](https://www.nature.com/articles/s43247-025-02424-9)
- **Colli Albani tephriphonolite** — [Fanesi et al. (2025)](https://www.sciencedirect.com/science/article/pii/S0377027325000125)
- **Anhydrous metaluminous/peralkaline haplogranite** — [Stopponi et al. (2026)](https://www.sciencedirect.com/science/article/pii/S0009254125005868)
- **Vesuvio phonotephrite (472 CE)** — [Dominijanni et al. (2026)](https://www.sciencedirect.com/science/article/pii/S0012821X25005126)
    """)

    # ── Model selection ───────────────────────────────────────────────────────
    comp_model = st.selectbox("Select composition model:", [
        "Stromboli basalt — Valdivia et al. (2023)",
        "Peridotite — Di Genova et al. (2023)",
        "Anhydrous andesite — Valdivia et al. (2025)",
        "Colli Albani tephriphonolite — Fanesi et al. (2025)",
        "Anhydrous metaluminous and peralkaline haplogranitic melts — Stopponi et al. (2026)",
        "Vesuvio phonotephrite (472 CE) — Dominijanni et al. (2026)",
    ])

    # ── Stromboli parameters (Valdivia et al. 2023, Table S1) ─────────────────
    # From supplementary MYEGA calculator Excel
    STRM = {
        'name':    'Stromboli basalt',
        'ref':     'Valdivia et al. (2023)',
        'A':       -2.93,
        'Tg_d':    940.1,    # K, anhydrous Tg
        'm':       40.7,     # anhydrous fragility index
        'b':       0.2351,   # Gordon-Taylor param
        'c':       1.234,
        'd':       -1.47,
        'Tg_H2O':  136.0,    # K, Tg of pure water
        'MW_dry':  64.0,     # g/mol effective molar weight of anhydrous melt
    }

    # ── Vesuvio phonotephrite parameters (Dominijanni et al. 2026) ──────────
    POX = {
        'name':    'Vesuvio phonotephrite (472 CE)',
        'ref':     'Dominijanni et al. (2026)',
        'A':       -2.93,
        'Tg_d':    917.05,   # K, anhydrous Tg
        'm_d':     35.92,    # anhydrous fragility index (H2O=0)
        'm_slope': 0.18857,  # dm/dx [mol%^-1] — linear fit (Excel cell D2)
        'b':       0.1322,   # Gordon-Taylor param (Excel cell E2)
        'c':       1.54232,  # Excel cell F2
        'd':       -1.40674, # Excel cell G2
        'Tg_H2O':  136.0,    # K (Excel cell H2)
        # Quadratic wt% → mol% conversion specific to this composition (Excel cell C8)
        # x_mol = mol_a * wt^2 + mol_b * wt
        'mol_a':   -0.0801,
        'mol_b':    3.6258,
        'm_mode':  'linear', # m(x) = m_d + m_slope * x_mol
        'MW_dry':  None,     # not used — quadratic formula in mol_a/mol_b
    }
    # Select active model
    # ── Peridotite parameters (Di Genova et al. 2023, Chem. Geol.) ──────────
    PRD = {
        'name':    'Peridotite melt',
        'ref':     'Di Genova et al. (2023)',
        'A':       -2.93,
        'Tg_d':    993.3,    # K
        'm_d':     59.069,   # anhydrous fragility
        'm_slope': -0.6281,  # dm/dx — NEGATIVE: m decreases with H2O
        'b':       0.2737,
        'c':       1.686,
        'd':       -1.779,
        'Tg_H2O':  136.0,
        'MW_dry':  51.01,    # g/mol — from S44F6 (Di Genova et al. 2023), using FeOtot
        'm_mode':  'linear',
    }

    # ── HPG8+Na parameters (Stopponi et al. 2026, Chem. Geol.) ─────────────
    # Model for anhydrous peralkaline rhyolite; Excess Na2O replaces water
    # All MYEGA params depend on Excess Na2O (mol%)
    HPG8_PARAMS = {
        'J5': -1.55, 'K5': 6.35, 'L5': 0.55,   # A = J5 - K5*L5^x
        'J7': 1034,  'K7': -0.1,                 # Tg = J7*(1+x)^K7
        'J9': 22.95, 'K9': 0.7,                  # m  = J9 + K9*x
    }
    HPG8 = {
        'name':      'HPG8+Na peralkaline rhyolite',
        'ref':       'Stopponi et al. (2026)',
        'model':     'HPG8',        # special model — not MYEGA with fixed A
        'x_label':   'Excess Na₂O (mol%)',
        'x_min':     0.0,
        'x_max':     20.0,
        'x_default': '0, 2, 5, 10, 15, 20',
        'params':    HPG8_PARAMS,
    }

    # ── Andesite parameters (Valdivia et al. 2025, Commun. Earth Environ.) ────
    # Anhydrous andesite; transition metals (FeOtot+TiO2+MnO) wt% relative to AND100
    AND100_comp = {'SiO2':60.3754,'TiO2':0.7937,'Al2O3':16.6937,'FeO':6.7673,
                   'MnO':0.1729,'MgO':3.0037,'CaO':6.6237,'Na2O':3.5041,'K2O':1.5756,'P2O5':0.1825}
    AND0_comp   = {'SiO2':65.9068,'TiO2':0,    'Al2O3':18.0141,'FeO':0,
                   'MnO':0,       'MgO':3.2123,'CaO':7.2568,'Na2O':3.8227,'K2O':1.7495,'P2O5':0}
    AND = {
        'name':      'Anhydrous andesite',
        'ref':       'Valdivia et al. (2025)',
        'model':     'AND',
        'A':         -2.93,
        'x_label':   'Transition metals content (wt% relative to AND100)',
        'x_min':     0.0,
        'x_max':     100.0,
        'x_default': '0, 25, 50, 75, 100',
        'AND100':    AND100_comp,
        'AND0':      AND0_comp,
    }

    # ── Colli Albani tephriphonolite (Fanesi et al. 2025, JVGR) ─────────────
    PNR = {
        'name':    'Colli Albani tephriphonolite',
        'ref':     'Fanesi et al. (2025)',
        'A':       -2.93,
        'Tg_d':    899.95,   # K, anhydrous Tg
        'm_d':     31.10585, # anhydrous fragility index (at H2O=0)
        'm_slope':  0.2709,  # dm/dx [mol%^-1] — positive: m increases with H2O
        'b':       0.25557,
        'c':       1.1576,
        'd':       -1.34585,
        'Tg_H2O':  136.0,    # K
        # Quadratic wt%→mol% conversion (Excel cell C8):
        # x_mol = mol_a*wt^2 + mol_b*wt + mol_c
        'mol_a':   -0.0854,
        'mol_b':    3.7366,
        'mol_c':    0.0105,
        'm_mode':  'linear',
    }

    ACTIVE = (HPG8 if 'Stopponi' in comp_model or 'haplogranitic' in comp_model
              else AND if '2025' in comp_model and 'Valdivia' in comp_model
              else PNR if 'Fanesi' in comp_model or 'Colli Albani' in comp_model
              else PRD if 'Di Genova' in comp_model
              else POX if 'Dominijanni' in comp_model or '472 CE' in comp_model
              else STRM)

    with st.expander("📋 Model parameters"):
        _P = ACTIVE
        if _P.get('model') in ('HPG8', 'AND'):
            if _P.get('model') == 'HPG8':
                st.markdown(f"""
| Parameter | Value | Description |
|-----------|-------|-------------|
| J5 | {_P['params']['J5']} | Constant in A(x) formula |
| K5 | {_P['params']['K5']} | Coefficient in A(x) formula |
| L5 | {_P['params']['L5']} | Base in A(x): A = J5 − K5·L5^x |
| J7 | {_P['params']['J7']} K | Constant in Tg(x): Tg = J7·(1+x)^K7 |
| K7 | {_P['params']['K7']} | Exponent in Tg(x) |
| J9 | {_P['params']['J9']} | Intercept in m(x): m = J9 + K9·x |
| K9 | {_P['params']['K9']} | Slope in m(x) |
| x range | 0 – 20 mol% | Calibrated range for Excess Na₂O |
                """)
            else:  # AND
                st.markdown("""
| Parameter | Value | Description |
|-----------|-------|-------------|
| A | −2.93 | Pre-exponential term [log Pa·s] |
| Tg(x) | 737 − 0.26077·x − 0.00569·x² (°C) | Quadratic in TM content x [0–100] |
| m(x) | 31.8 − 0.013·x | Linear in TM content x |
| x | TM content wt% relative to AND100 | FeOtot + TiO₂ + MnO |
| x range | 0 – 100 % | 0 = no TM; 100 = AND100 composition |
                """)
        else:
            _mode = _P.get('m_mode', 'constant')
            if _mode == 'linear':
                _m_label = f"{_P['m_d']} + {_P['m_slope']} · x_mol — linear calibration"
                _mol_conv = (f"x_mol = {_P['mol_a']}·wt² + {_P['mol_b']}·wt  (empirical fit)"
                             if 'mol_a' in _P else f"standard formula using MW_dry = {_P.get('MW_dry','')}")
            else:
                _m_label = f"{_P.get('m', _P.get('m_d', ''))} (constant)"
                _mol_conv = f"standard formula using MW_dry = {_P.get('MW_dry','')}"
            st.markdown(f"""
| Parameter | Value | Description |
|-----------|-------|-------------|
| A | {_P['A']} | Pre-exponential term [log Pa·s] |
| Tg_dry | {_P['Tg_d']} K  ({_P['Tg_d']-273.15:.1f} °C) | Anhydrous glass transition temperature |
| m | {_m_label} | Fragility index |
| b | {_P['b']} | Gordon-Taylor mixing parameter |
| c | {_P['c']} | Excess mixing term |
| d | {_P['d']} | Higher-order correction |
| Tg_H₂O | {_P['Tg_H2O']} K | Glass transition of pure water |
| wt% → mol% | {_mol_conv} | Conversion formula for H₂O content |
            """)

    # ── Helper functions (generic, work for any model) ───────────────────────
    def strm_wt_to_mol(h2o_wt, MW_dry):
        """Generic wt% → mol% conversion using molar weight of dry melt."""
        n_h2o = h2o_wt / 18.015
        n_dry = (100.0 - h2o_wt) / MW_dry
        return n_h2o / (n_h2o + n_dry) * 100.0 if (n_h2o + n_dry) > 0 else 0.0

    def wt_to_mol(h2o_wt, p):
        """Wt% → mol% using model-specific method.
        If p has mol_a/mol_b: quadratic empirical fit (e.g. Pollena Excel cell C8).
        Otherwise: generic formula using MW_dry.
        """
        if 'mol_a' in p:
            return p['mol_a'] * h2o_wt**2 + p['mol_b'] * h2o_wt + p.get('mol_c', 0.0)
        mw = p.get('MW_dry')
        if mw:
            return strm_wt_to_mol(h2o_wt, mw)
        return h2o_wt  # fallback

    def strm_tg(x_mol_pct, p):
        """Tg(H2O) — Eqs. 9-10, Langhammer et al. (2021)."""
        x = x_mol_pct / 100.0
        denom = p['b'] * (1.0 - x) + x
        if denom == 0: return p['Tg_d']
        w1 = x / denom
        w2 = p['b'] * (1.0 - x) / denom
        return (w1 * p['Tg_H2O'] + w2 * p['Tg_d']
                + p['c'] * w1 * w2 * (p['Tg_d'] - p['Tg_H2O'])
                + p['d'] * w1 * w2**2 * (p['Tg_d'] - p['Tg_H2O']))

    def get_m(x_mol_pct, p):
        """Fragility index:
          - 'constant' (Stromboli): m = m_d fixed
          - 'linear'   (Pollena):   m = m_d + m_slope * x  (x in mol%)
        """
        mode = p.get('m_mode', 'constant')
        if mode == 'linear':
            return p['m_d'] + p['m_slope'] * x_mol_pct
        else:
            return p.get('m', p.get('m_d', 35.0))

    def strm_myega(T_K, Tg_K, m, A):
        """MYEGA viscosity equation — Mauro et al. (2009)."""
        ratio = Tg_K / T_K
        return A + (12.0 - A) * ratio * np.exp((m / (12.0 - A) - 1.0) * (ratio - 1.0))


    # ── HPG8 model: special rendering ────────────────────────────────────────
    if ACTIVE.get('model') == 'HPG8':
        P = ACTIVE['params']

        def hpg8_A(x):  return P['J5'] - P['K5'] * P['L5']**x
        def hpg8_Tg(x): return P['J7'] * (1+x)**P['K7']
        def hpg8_m(x):  return P['J9'] + P['K9'] * x
        def hpg8_visc(T_K, x):
            A_x=hpg8_A(x); Tg_x=hpg8_Tg(x); m_x=hpg8_m(x); r=Tg_x/T_K
            return A_x+(12-A_x)*r*np.exp((m_x/(12-A_x)-1)*(r-1))

        st.subheader("\u2699\ufe0f Parameters")
        hc1,hc2,hc3,hc4 = st.columns(4)
        with hc1:
            hT_min = st.number_input("T min (\u00b0C):", value=400.0, step=50.0, min_value=300.0, max_value=1600.0, key='hpg_Tmin')
        with hc2:
            hT_max = st.number_input("T max (\u00b0C):", value=1200.0, step=50.0, min_value=300.0, max_value=1600.0, key='hpg_Tmax')
        with hc3:
            hx_input = st.text_input("Excess Na\u2082O (mol%, 0\u201320, comma-separated):", value=ACTIVE['x_default'], key='hpg_x')
            try:
                hx_list = sorted(set([max(0.0,min(20.0,float(v.strip()))) for v in hx_input.split(',') if v.strip()]))
            except:
                hx_list = [0,5,10,15,20]
        with hc4:
            hT_fixed = st.number_input("Fixed T for \u03b7 vs Excess Na\u2082O (\u00b0C):", value=800.0, step=50.0, min_value=300.0, max_value=1600.0, key='hpg_Tf')

        if hT_max <= hT_min: st.error("T max must be greater than T min."); st.stop()

        hT_arr=np.linspace(hT_min+273.15,hT_max+273.15,300)
        hx_dense=np.linspace(0,20,200)
        hcmap=plt.get_cmap('plasma',max(len(hx_list),2))

        hresults=[{'x':x,'A':round(hpg8_A(x),4),'Tg_K':round(hpg8_Tg(x),2),
                   'Tg_C':round(hpg8_Tg(x)-273.15,2),'m':round(hpg8_m(x),3)} for x in hx_list]

        hfig,(hax1,hax2,hax3)=plt.subplots(1,3,figsize=(18,5))

        for i,x in enumerate(hx_list):
            hax1.plot(hT_arr-273.15,[hpg8_visc(T,x) for T in hT_arr],
                      color=hcmap(i),linewidth=2,label=f"{x:.0f} mol%")
        hax1.set_xlabel("Temperature (\u00b0C)",fontsize=11)
        hax1.set_ylabel("log\u2081\u2080(\u03b7 / Pa\u00b7s)",fontsize=11)
        hax1.set_title("Viscosity vs Temperature\nHPG8+Na (Stopponi et al. 2026)",fontsize=11,fontweight='bold')
        hax1.legend(fontsize=8,loc='upper right',title='Excess Na\u2082O',title_fontsize=8)
        hax1.grid(True,linestyle='--',alpha=0.4)

        hax2b=hax2.twinx()
        hl1,=hax2.plot(hx_dense,[hpg8_Tg(x)-273.15 for x in hx_dense],'steelblue',linewidth=2.5,label='Tg (\u00b0C)')
        hax2.scatter([r['x'] for r in hresults],[r['Tg_C'] for r in hresults],
                     color=[hcmap(i) for i in range(len(hresults))],s=60,zorder=5,edgecolors='black',linewidths=0.8)
        hl2,=hax2b.plot(hx_dense,[hpg8_m(x) for x in hx_dense],'tomato',linewidth=2.5,linestyle='--',label='m')
        hax2b.scatter([r['x'] for r in hresults],[r['m'] for r in hresults],
                      color=[hcmap(i) for i in range(len(hresults))],s=50,zorder=5,edgecolors='black',linewidths=0.8,marker='D')
        hax2.set_xlabel("Excess Na\u2082O (mol%)",fontsize=11)
        hax2.set_ylabel("Tg (\u00b0C)",fontsize=11,color='steelblue')
        hax2b.set_ylabel("Fragility index m",fontsize=11,color='tomato')
        hax2.tick_params(axis='y',labelcolor='steelblue'); hax2b.tick_params(axis='y',labelcolor='tomato')
        hax2.set_title("Tg and m vs Excess Na\u2082O\nHPG8+Na (Stopponi et al. 2026)",fontsize=11,fontweight='bold')
        hax2.legend(handles=[hl1,hl2],fontsize=8,loc='center right')
        hax2.grid(True,linestyle='--',alpha=0.4)

        hax3.plot(hx_dense,[hpg8_visc(hT_fixed+273.15,x) for x in hx_dense],'purple',linewidth=2.5)
        hax3.scatter(hx_list,[hpg8_visc(hT_fixed+273.15,x) for x in hx_list],
                     color=[hcmap(i) for i in range(len(hx_list))],s=70,zorder=5,edgecolors='black',linewidths=0.8)
        hax3.set_xlabel("Excess Na\u2082O (mol%)",fontsize=11)
        hax3.set_ylabel("log\u2081\u2080(\u03b7 / Pa\u00b7s)",fontsize=11)
        hax3.set_title(f"Viscosity vs Excess Na\u2082O at {hT_fixed:.0f} \u00b0C\nHPG8+Na (Stopponi et al. 2026)",fontsize=11,fontweight='bold')
        hax3.grid(True,linestyle='--',alpha=0.4)

        plt.tight_layout()
        st.pyplot(hfig)

        # Save panels using axis bounding boxes — no new figures needed
        hbuf_all = io.BytesIO()
        hfig.savefig(hbuf_all, format='png', dpi=200, bbox_inches='tight'); hbuf_all.seek(0)
        def _bb(fig, ax):
            _buf = io.BytesIO()
            bb = ax.get_tightbbox(fig.canvas.get_renderer()).transformed(fig.dpi_scale_trans.inverted())
            fig.savefig(_buf, format='png', dpi=200, bbox_inches=bb); _buf.seek(0); return _buf
        _hb1=_bb(hfig,hax1); _hb2=_bb(hfig,hax2); _hb3=_bb(hfig,hax3)
        plt.close(hfig)
        _hcols=st.columns(4)
        with _hcols[0]: st.download_button('⬇️ All figures',data=hbuf_all,file_name='HPG8_all.png',mime='image/png',key='dl_hpg_all')
        with _hcols[1]: st.download_button('⬇️ Viscosity vs T',data=_hb1,file_name='HPG8_visc_vs_T.png',mime='image/png',key='dl_hpg_p0')
        with _hcols[2]: st.download_button('⬇️ Tg & m',data=_hb2,file_name='HPG8_Tg_m.png',mime='image/png',key='dl_hpg_p1')
        with _hcols[3]: st.download_button('⬇️ η vs Excess Na₂O',data=_hb3,file_name='HPG8_visc_vs_Na2O.png',mime='image/png',key='dl_hpg_p2')

        st.subheader("\U0001f4ca Model parameters summary")
        st.dataframe(pd.DataFrame(hresults).rename(columns={'x':'Excess Na\u2082O (mol%)','A':'log10\u03b7\u221e','Tg_K':'Tg (K)','Tg_C':'Tg (\u00b0C)','m':'m'}),use_container_width=True,hide_index=True)

        st.subheader("\U0001f321\ufe0f Viscosity at specific conditions")
        hc1b,hc2b=st.columns(2)
        with hc1b: hT_sp=st.number_input("T (\u00b0C):",value=800.0,step=50.0,min_value=300.0,max_value=1600.0,key='hpg_Tsp')
        with hc2b: hx_sp=st.number_input("Excess Na\u2082O (mol%, 0\u201320):",value=10.0,step=1.0,min_value=0.0,max_value=20.0,key='hpg_xsp')
        hlv_sp=hpg8_visc(hT_sp+273.15,hx_sp)
        st.metric("log\u2081\u2080(\u03b7 / Pa\u00b7s)",f"{hlv_sp:.3f}",help=f"A={hpg8_A(hx_sp):.4f} | Tg={hpg8_Tg(hx_sp)-273.15:.1f}\u00b0C | m={hpg8_m(hx_sp):.3f}")

        hbuf_xl=io.BytesIO()
        hrows=[{'Excess Na2O (mol%)':x,'T (\u00b0C)':Tc,'T (K)':Tc+273.15,
                'log10_visc':round(hpg8_visc(Tc+273.15,x),4),'A':round(hpg8_A(x),4),
                'Tg (K)':round(hpg8_Tg(x),2),'m':round(hpg8_m(x),3)}
               for x in hx_list for Tc in np.arange(hT_min,hT_max+25,25)]
        pd.DataFrame(hrows).to_excel(hbuf_xl,index=False,sheet_name='HPG8_Na_viscosity'); hbuf_xl.seek(0)
        st.download_button("\u2b07\ufe0f Download Excel",data=hbuf_xl,file_name="HPG8_Na_viscosity.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",key="dl_hpg_xl")
        st.caption("\U0001f4d6 MYEGA: [Mauro et al. (2009)](https://www.pnas.org/doi/10.1073/pnas.0911705106), *PNAS* 106, 19780\u201319784. "
                   "HPG8+Na: [Stopponi et al. (2026)](https://www.sciencedirect.com/science/article/pii/S0009254125005868), *Chem. Geol.* "
                   "Model calibrated for Excess Na\u2082O = 0\u201320 mol%.")

    elif ACTIVE.get('model') == 'AND':
        P = ACTIVE

        def and_Tg(x):  return 737 + (-0.26077)*x + (-0.00569)*x**2   # °C
        def and_m(x):   return -0.013*x + 31.8
        def and_visc(T_K, x):
            Tg_K = and_Tg(x) + 273.15
            m_x  = and_m(x)
            r    = Tg_K / T_K
            return P['A'] + (12-P['A'])*r*np.exp((m_x/(12-P['A'])-1)*(r-1))

        def and_comp(x):
            return {k: (x/100)*P['AND100'][k] + (1-x/100)*P['AND0'].get(k,0) for k in P['AND100']}

        st.subheader("⚙️ Parameters")
        ac1,ac2,ac3,ac4 = st.columns(4)
        with ac1:
            aT_min = st.number_input("T min (°C):", value=700.0, step=50.0, min_value=300.0, max_value=1600.0, key='and_Tmin')
        with ac2:
            aT_max = st.number_input("T max (°C):", value=1300.0, step=50.0, min_value=300.0, max_value=1600.0, key='and_Tmax')
        with ac3:
            ax_input = st.text_input("TM content (%, 0–100, comma-separated):", value=P['x_default'], key='and_x')
            try:
                ax_list = sorted(set([max(0.0,min(100.0,float(v.strip()))) for v in ax_input.split(',') if v.strip()]))
            except:
                ax_list = [0, 25, 50, 75, 100]
        with ac4:
            aT_fixed = st.number_input("Fixed T for η vs TM content (°C):", value=1000.0, step=50.0,
                                        min_value=300.0, max_value=1600.0, key='and_Tf')

        if aT_max <= aT_min: st.error("T max must be greater than T min."); st.stop()

        aT_arr   = np.linspace(aT_min+273.15, aT_max+273.15, 300)
        ax_dense = np.linspace(0, 100, 200)
        acmap    = plt.get_cmap('viridis', max(len(ax_list), 2))

        aresults = [{'x':x,'Tg_C':round(and_Tg(x),2),'Tg_K':round(and_Tg(x)+273.15,2),'m':round(and_m(x),3)} for x in ax_list]

        afig, (aax1, aax2, aax3) = plt.subplots(1, 3, figsize=(18,5))

        # Panel 1: viscosity vs T
        for i,x in enumerate(ax_list):
            aax1.plot(aT_arr-273.15, [and_visc(T,x) for T in aT_arr],
                      color=acmap(i), linewidth=2, label=f"{x:.0f}%")
        aax1.set_xlabel("Temperature (°C)", fontsize=11)
        aax1.set_ylabel("log₁₀(η / Pa·s)", fontsize=11)
        aax1.set_title("Viscosity vs Temperature\nAndesite (Valdivia et al. 2025)", fontsize=11, fontweight='bold')
        aax1.legend(fontsize=8, loc='upper right', title='TM content', title_fontsize=8)
        aax1.grid(True, linestyle='--', alpha=0.4)

        # Panel 2: Tg and m vs TM content
        aax2b = aax2.twinx()
        al1, = aax2.plot(ax_dense, [and_Tg(x) for x in ax_dense], 'steelblue', linewidth=2.5, label='Tg (°C)')
        aax2.scatter([r['x'] for r in aresults], [r['Tg_C'] for r in aresults],
                     color=[acmap(i) for i in range(len(aresults))], s=60, zorder=5, edgecolors='black', linewidths=0.8)
        al2, = aax2b.plot(ax_dense, [and_m(x) for x in ax_dense], 'tomato', linewidth=2.5, linestyle='--', label='m')
        aax2b.scatter([r['x'] for r in aresults], [r['m'] for r in aresults],
                      color=[acmap(i) for i in range(len(aresults))], s=50, zorder=5, edgecolors='black', linewidths=0.8, marker='D')
        aax2.set_xlabel("TM content (wt% relative to AND100)", fontsize=11)
        aax2.set_ylabel("Tg (°C)", fontsize=11, color='steelblue')
        aax2b.set_ylabel("Fragility index m", fontsize=11, color='tomato')
        aax2.tick_params(axis='y', labelcolor='steelblue'); aax2b.tick_params(axis='y', labelcolor='tomato')
        aax2.set_title("Tg and m vs TM content\nAndesite (Valdivia et al. 2025)", fontsize=11, fontweight='bold')
        aax2.legend(handles=[al1,al2], fontsize=8, loc='center right')
        aax2.grid(True, linestyle='--', alpha=0.4)

        # Panel 3: viscosity vs TM content at fixed T
        aax3.plot(ax_dense, [and_visc(aT_fixed+273.15,x) for x in ax_dense], 'purple', linewidth=2.5)
        aax3.scatter(ax_list, [and_visc(aT_fixed+273.15,x) for x in ax_list],
                     color=[acmap(i) for i in range(len(ax_list))], s=70, zorder=5, edgecolors='black', linewidths=0.8)
        aax3.set_xlabel("TM content (wt% relative to AND100)", fontsize=11)
        aax3.set_ylabel("log₁₀(η / Pa·s)", fontsize=11)
        aax3.set_title(f"Viscosity vs TM content at {aT_fixed:.0f} °C\nAndesite (Valdivia et al. 2025)", fontsize=11, fontweight='bold')
        aax3.grid(True, linestyle='--', alpha=0.4)

        plt.tight_layout()
        st.pyplot(afig)

        # Save panels using bounding boxes
        abuf_all = io.BytesIO()
        afig.savefig(abuf_all, format='png', dpi=200, bbox_inches='tight'); abuf_all.seek(0)
        def _bba(ax):
            _buf=io.BytesIO()
            bb=ax.get_tightbbox(afig.canvas.get_renderer()).transformed(afig.dpi_scale_trans.inverted())
            afig.savefig(_buf,format='png',dpi=200,bbox_inches=bb); _buf.seek(0); return _buf
        _ab1=_bba(aax1); _ab2=_bba(aax2); _ab3=_bba(aax3)
        plt.close(afig)
        _acols = st.columns(4)
        with _acols[0]: st.download_button('⬇️ All figures', data=abuf_all, file_name='Andesite_all.png', mime='image/png', key='dl_and_all')
        with _acols[1]: st.download_button('⬇️ Viscosity vs T', data=_ab1, file_name='Andesite_visc_vs_T.png', mime='image/png', key='dl_and_p0')
        with _acols[2]: st.download_button('⬇️ Tg & m', data=_ab2, file_name='Andesite_Tg_m.png', mime='image/png', key='dl_and_p1')
        with _acols[3]: st.download_button('⬇️ η vs TM', data=_ab3, file_name='Andesite_visc_vs_TM.png', mime='image/png', key='dl_and_p2')

        st.subheader("📊 Model summary")
        st.dataframe(pd.DataFrame(aresults).rename(columns={'x':'TM content (%)','Tg_C':'Tg (°C)','Tg_K':'Tg (K)','m':'m'}),
                     use_container_width=True, hide_index=True)

        st.subheader("🌡️ Viscosity at specific conditions")
        ac1b,ac2b = st.columns(2)
        with ac1b: aT_sp=st.number_input("T (°C):",value=1000.0,step=50.0,min_value=300.0,max_value=1600.0,key='and_Tsp')
        with ac2b: ax_sp=st.number_input("TM content (%, 0–100):",value=50.0,step=5.0,min_value=0.0,max_value=100.0,key='and_xsp')
        alv_sp=and_visc(aT_sp+273.15,ax_sp)
        st.metric("log₁₀(η / Pa·s)",f"{alv_sp:.3f}",help=f"Tg={and_Tg(ax_sp):.1f}°C | m={and_m(ax_sp):.3f}")

        abuf_xl=io.BytesIO()
        arows=[{'TM content (%)':x,'T (°C)':Tc,'T (K)':Tc+273.15,
                'log10_visc':round(and_visc(Tc+273.15,x),4),
                'Tg (°C)':round(and_Tg(x),2),'m':round(and_m(x),3)}
               for x in ax_list for Tc in np.arange(aT_min,aT_max+25,25)]
        pd.DataFrame(arows).to_excel(abuf_xl,index=False,sheet_name='Andesite_viscosity'); abuf_xl.seek(0)
        st.download_button('⬇️ Download Excel',data=abuf_xl,file_name='Andesite_viscosity.xlsx',
                           mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',key='dl_and_xl')
        st.caption("📖 MYEGA: [Mauro et al. (2009)](https://www.pnas.org/doi/10.1073/pnas.0911705106), *PNAS* 106, 19780–19784. "
                   "Andesite model: [Valdivia et al. (2025)](https://www.nature.com/articles/s43247-025-02424-9), "
                   "*Commun. Earth Environ.* AND100 = reference andesite; TM = FeOtot+TiO2+MnO.")

    else:
        # ── Inputs layout: H2O full width, T controls above their panels ─────────
        st.subheader("⚙️ Parameters")

        # H2O full width (spans all 3 panels)
        h2o_input_s = st.text_input("H₂O contents (wt%, comma-separated):",
                                     value="0, 0.5, 1, 2, 3, 4, 5")
        try:
            h2o_list_s = sorted(set([float(x.strip()) for x in h2o_input_s.split(',') if x.strip()]))
        except:
            h2o_list_s = [0.0, 1.0, 2.0, 3.0]
        if 0.0 not in h2o_list_s:
            h2o_list_s = [0.0] + h2o_list_s

        # T min/max above panel 1 | blank above panel 2 | Fixed T above panel 3
        _pc1, _pc2, _pc3 = st.columns(3)
        with _pc1:
            T_min_c = st.number_input("T min (°C):", value=700.0, step=50.0,
                                       min_value=300.0, max_value=1600.0)
            T_max_c = st.number_input("T max (°C):", value=1300.0, step=50.0,
                                       min_value=300.0, max_value=1600.0)
        with _pc2:
            pass  # no input needed for Tg/m panel
        with _pc3:
            T_fixed_c = st.number_input("Fixed T for η vs H₂O (°C):", value=1050.0,
                                         step=50.0, min_value=300.0, max_value=1600.0,
                                         key='strm_T_fixed')

        if T_max_c <= T_min_c:
            st.error("T max must be greater than T min.")
            st.stop()

        T_arr  = np.linspace(T_min_c + 273.15, T_max_c + 273.15, 300)
        cmap_s = plt.get_cmap('plasma', len(h2o_list_s))


        # Compute all data
        results_s = []
        for i_h, h2o_wt in enumerate(h2o_list_s):
            x_mol = wt_to_mol(h2o_wt, ACTIVE)
            Tg_h  = strm_tg(x_mol, ACTIVE)
            results_s.append({'h2o_wt': h2o_wt, 'x_mol': round(x_mol, 3),
                              'Tg_K': round(Tg_h, 2), 'Tg_C': round(Tg_h - 273.15, 2)})
        h2o_dense = np.linspace(0, max(max(h2o_list_s), 5.0) * 1.05, 200)
        tg_curve  = [strm_tg(wt_to_mol(h, ACTIVE), ACTIVE) - 273.15 for h in h2o_dense]
        h2o_vh    = np.linspace(0, max(max(h2o_list_s), 5.0), 200)
        visc_vh   = [strm_myega(T_fixed_c+273.15,
                                 strm_tg(wt_to_mol(h, ACTIVE), ACTIVE),
                                 get_m(wt_to_mol(h, ACTIVE), ACTIVE), ACTIVE['A']) for h in h2o_vh]
        visc_pts  = [strm_myega(T_fixed_c+273.15,
                                 strm_tg(wt_to_mol(h, ACTIVE), ACTIVE),
                                 get_m(wt_to_mol(h, ACTIVE), ACTIVE), ACTIVE['A']) for h in h2o_list_s]

        # Single figure — 3 panels on one row
        fig_s, (ax_visc, ax_tgm, ax_vh2) = plt.subplots(1, 3, figsize=(18, 5))

        # Panel 1: viscosity vs T
        _has_linear_m = ACTIVE.get('m_mode', 'constant') != 'constant'
        for i_h, h2o_wt in enumerate(h2o_list_s):
            x_mol = wt_to_mol(h2o_wt, ACTIVE)
            Tg_h  = strm_tg(x_mol, ACTIVE)
            _m_h  = get_m(x_mol, ACTIVE)
            visc  = [strm_myega(T, Tg_h, _m_h, ACTIVE['A']) for T in T_arr]
            _lbl  = f"{h2o_wt:.1f} wt%" + (" (m lin.)" if _has_linear_m else "")
            ax_visc.plot(T_arr - 273.15, visc, color=cmap_s(i_h), linewidth=2, label=_lbl)
            if _has_linear_m:
                visc_mc = [strm_myega(T, Tg_h, ACTIVE['m_d'], ACTIVE['A']) for T in T_arr]
                ax_visc.plot(T_arr - 273.15, visc_mc, color=cmap_s(i_h), linewidth=1.5,
                             linestyle=':', alpha=0.7,
                             label=f"{h2o_wt:.1f} wt% (m const.)" if i_h == 0 else "_")
        ax_visc.set_xlabel("Temperature (°C)", fontsize=11)
        ax_visc.set_ylabel("log₁₀(η / Pa·s)", fontsize=11)
        ax_visc.set_title(f"Viscosity vs Temperature\n{ACTIVE['name']}",
                           fontsize=11, fontweight='bold')
        ax_visc.legend(fontsize=7, loc='upper right')
        ax_visc.grid(True, linestyle='--', alpha=0.4)

        # Panel 2: Tg and m vs H2O (twin axes, m constant)
        ax_tgm2 = ax_tgm.twinx()
        l1, = ax_tgm.plot(h2o_dense, tg_curve, 'steelblue', linewidth=2.5, label='Tg (\u00b0C)')
        ax_tgm.scatter([r['h2o_wt'] for r in results_s], [r['Tg_C'] for r in results_s],
                       color=[cmap_s(i) for i in range(len(results_s))],
                       s=60, zorder=5, edgecolors='black', linewidths=0.8)
        _m_d_val = ACTIVE.get('m_d', ACTIVE.get('m', 35.0))
        if _has_linear_m:
            m_curve_dense = [get_m(wt_to_mol(h, ACTIVE), ACTIVE) for h in h2o_dense]
            l2, = ax_tgm2.plot(h2o_dense, m_curve_dense, color='tomato', linewidth=2.5,
                                linestyle='--', label=f"m linear: {ACTIVE['m_d']} + {ACTIVE['m_slope']:.5f}·x")
            _m_pts = [get_m(wt_to_mol(r['h2o_wt'], ACTIVE), ACTIVE) for r in results_s]
            ax_tgm2.scatter([r['h2o_wt'] for r in results_s], _m_pts,
                            color=[cmap_s(i) for i in range(len(results_s))],
                            s=50, zorder=5, edgecolors='black', linewidths=0.7, marker='D')
            l2b, = ax_tgm2.plot([h2o_dense[0], h2o_dense[-1]], [_m_d_val, _m_d_val],
                                 color='darkorange', linewidth=1.8, linestyle=':',
                                 label=f"m constant = {_m_d_val}")
            _legend_handles = [l1, l2, l2b]
        else:
            l2, = ax_tgm2.plot([h2o_dense[0], h2o_dense[-1]], [_m_d_val, _m_d_val],
                                color='tomato', linewidth=2.5, linestyle='--',
                                label=f"m = {_m_d_val} (constant)")
            _legend_handles = [l1, l2]
        ax_tgm2.set_ylim(min(_m_d_val*0.85, 20), max(_m_d_val*1.15, 50))
        ax_tgm.set_xlabel("H\u2082O (wt%)", fontsize=11)
        ax_tgm.set_ylabel("Tg (\u00b0C)", fontsize=11, color='steelblue')
        ax_tgm2.set_ylabel("Fragility index m", fontsize=11, color='tomato')
        ax_tgm.tick_params(axis='y', labelcolor='steelblue')
        ax_tgm2.tick_params(axis='y', labelcolor='tomato')
        ax_tgm.set_title(f"Tg and m vs H\u2082O\n{ACTIVE['name']}",
                          fontsize=11, fontweight='bold')
        ax_tgm.legend(handles=_legend_handles, fontsize=8, loc='lower left')
        ax_tgm.grid(True, linestyle='--', alpha=0.4)

        # Panel 3: viscosity vs H2O at fixed T
        _lbl_vh = 'm linear' if _has_linear_m else 'm'
        ax_vh2.plot(h2o_vh, visc_vh, 'purple', linewidth=2.5, label=_lbl_vh)
        ax_vh2.scatter(h2o_list_s, visc_pts,
                       color=[cmap_s(i) for i in range(len(h2o_list_s))],
                       s=70, zorder=5, edgecolors='black', linewidths=0.8)
        if _has_linear_m:
            visc_vh_mc = [strm_myega(T_fixed_c+273.15,
                                      strm_tg(wt_to_mol(h, ACTIVE), ACTIVE),
                                      ACTIVE['m_d'], ACTIVE['A']) for h in h2o_vh]
            ax_vh2.plot(h2o_vh, visc_vh_mc, 'darkorange', linewidth=1.8,
                        linestyle=':', label=f"m constant = {ACTIVE['m_d']}")
            ax_vh2.legend(fontsize=8, loc='upper right')
        ax_vh2.set_xlabel("H\u2082O (wt%)", fontsize=11)
        ax_vh2.set_ylabel("log\u2081\u2080(\u03b7 / Pa\u00b7s)", fontsize=11)
        ax_vh2.set_title(f"Viscosity vs H\u2082O at {T_fixed_c:.0f} \u00b0C\n{ACTIVE['name']}",
                         fontsize=11, fontweight='bold')
        ax_vh2.grid(True, linestyle='--', alpha=0.4)

        plt.tight_layout()
        st.pyplot(fig_s)

        # Save panels using axis bounding boxes — no new figures needed
        buf_fig_s = io.BytesIO()
        fig_s.savefig(buf_fig_s, format='png', dpi=200, bbox_inches='tight'); buf_fig_s.seek(0)
        def _bbg(ax):
            _buf = io.BytesIO()
            bb = ax.get_tightbbox(fig_s.canvas.get_renderer()).transformed(fig_s.dpi_scale_trans.inverted())
            fig_s.savefig(_buf, format='png', dpi=200, bbox_inches=bb); _buf.seek(0); return _buf
        _bp0=_bbg(ax_visc); _bp1=_bbg(ax_tgm); _bp2=_bbg(ax_vh2)
        plt.close(fig_s)
        _sname=ACTIVE['name'].replace(' ','_')
        _scols=st.columns(4)
        with _scols[0]: st.download_button('⬇️ All figures',data=buf_fig_s,file_name=f'{_sname}_all.png',mime='image/png',key='dl_strm_all')
        with _scols[1]: st.download_button('⬇️ Viscosity vs T',data=_bp0,file_name=f'{_sname}_visc_vs_T.png',mime='image/png',key='dl_strm_p0')
        with _scols[2]: st.download_button('⬇️ Tg & m vs H₂O',data=_bp1,file_name=f'{_sname}_Tg_m.png',mime='image/png',key='dl_strm_p1')
        with _scols[3]: st.download_button('⬇️ η vs H₂O',data=_bp2,file_name=f'{_sname}_visc_vs_H2O.png',mime='image/png',key='dl_strm_p2')

            # ── Summary table ─────────────────────────────────────────────────────────
        st.subheader("📊 Tg summary")
        df_tg = pd.DataFrame(results_s)
        df_tg.columns = ['H₂O (wt%)', 'H₂O (mol%)', 'Tg (K)', 'Tg (°C)']
        st.dataframe(df_tg, use_container_width=True, hide_index=True)

        # ── Viscosity at specific T and H2O ───────────────────────────────────────
        st.subheader("🌡️ Viscosity at specific conditions")
        col_sp1, col_sp2 = st.columns(2)
        with col_sp1:
            T_spec = st.number_input("Temperature (°C):", value=1100.0, step=50.0,
                                      min_value=300.0, max_value=1600.0, key='strm_T_spec')
        with col_sp2:
            h2o_spec = st.number_input("H₂O (wt%):", value=1.0, step=0.5,
                                        min_value=0.0, max_value=10.0, key='strm_h2o_spec')
        x_mol_sp   = wt_to_mol(h2o_spec, ACTIVE)
        Tg_sp      = strm_tg(x_mol_sp, ACTIVE)
        m_sp_var   = get_m(x_mol_sp, ACTIVE)           # linear or constant
        m_sp_cst   = ACTIVE.get('m_d', ACTIVE.get('m', 35.0))  # always constant
        lv_var     = strm_myega(T_spec + 273.15, Tg_sp, m_sp_var, ACTIVE['A'])
        lv_cst     = strm_myega(T_spec + 273.15, Tg_sp, m_sp_cst, ACTIVE['A'])
        _has_lin   = ACTIVE.get('m_mode', 'constant') != 'constant'

        if _has_lin:
            col_v1, col_v2 = st.columns(2)
            with col_v1:
                st.metric("log₁₀(η / Pa·s) — m linear",  f"{lv_var:.3f}",
                          help=f"m = {m_sp_var:.3f} | H₂O = {x_mol_sp:.2f} mol% | Tg = {Tg_sp-273.15:.1f} °C")
            with col_v2:
                st.metric("log₁₀(η / Pa·s) — m constant", f"{lv_cst:.3f}",
                          help=f"m = {m_sp_cst:.2f} | H₂O = {x_mol_sp:.2f} mol% | Tg = {Tg_sp-273.15:.1f} °C")
        else:
            visc_pa = 10**lv_var
            st.metric("log₁₀(η / Pa·s)", f"{lv_var:.3f}",
                      help=f"m = {m_sp_var:.2f} | H₂O = {x_mol_sp:.2f} mol% | Tg = {Tg_sp-273.15:.1f} °C")

        # ── Download ──────────────────────────────────────────────────────────────
        st.subheader("📥 Download results")
        buf_s = io.BytesIO()
        T_step_dl = 25.0
        T_dl = np.arange(T_min_c, T_max_c + T_step_dl, T_step_dl)
        rows_dl = []
        for h2o_wt in h2o_list_s:
            x_mol = wt_to_mol(h2o_wt, ACTIVE)
            Tg_h  = strm_tg(x_mol, ACTIVE)
            for Tc in T_dl:
                lv = strm_myega(Tc + 273.15, Tg_h, ACTIVE.get('m_d', ACTIVE.get('m', 35.0)), ACTIVE['A'])
                rows_dl.append({'H2O (wt%)': h2o_wt, 'H2O (mol%)': round(x_mol, 3),
                                'T (°C)': Tc, 'T (K)': Tc + 273.15,
                                'log10_visc': round(lv, 4)})
        pd.DataFrame(rows_dl).to_excel(buf_s, index=False, sheet_name='Stromboli_viscosity')
        buf_s.seek(0)
        st.download_button("⬇️ Download Excel", data=buf_s,
                           file_name=f"{ACTIVE['name'].replace(' ', '_')}_viscosity.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_strm_xl")

    if 'Fanesi' in comp_model or 'Colli Albani' in comp_model:
        st.caption("📖 MYEGA: [Mauro et al. (2009)](https://www.pnas.org/doi/10.1073/pnas.0911705106), *PNAS* 106, 19780\u201319784. "
                   "Colli Albani: [Fanesi et al. (2025)](https://www.sciencedirect.com/science/article/pii/S0377027325000125), *J. Volcanol. Geotherm. Res.* "
                   "Tg(H\u2082O): [Langhammer et al. (2021)](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918), *GGG* 22, e2021GC009918.")
    elif 'Di Genova' in comp_model:
        st.caption("📖 MYEGA: [Mauro et al. (2009)](https://www.pnas.org/doi/10.1073/pnas.0911705106), *PNAS* 106, 19780\u201319784. "
                   "Peridotite: [Di Genova et al. (2023)](https://www.sciencedirect.com/science/article/pii/S0009254123001407), *Chem. Geol.* "
                   "Tg(H\u2082O): [Langhammer et al. (2021)](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918), *GGG* 22, e2021GC009918.")
    elif '472 CE' in comp_model:
        st.caption("📖 MYEGA: [Mauro et al. (2009)](https://www.pnas.org/doi/10.1073/pnas.0911705106), *PNAS* 106, 19780\u201319784. "
                   "Vesuvio: [Dominijanni et al. (2026)](https://www.sciencedirect.com/science/article/pii/S0012821X25005126), *Earth Planet. Sci. Lett.* "
                   "Tg(H\u2082O): [Langhammer et al. (2021)](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918), *GGG* 22, e2021GC009918.")
    else:
        st.caption("📖 MYEGA: [Mauro et al. (2009)](https://www.pnas.org/doi/10.1073/pnas.0911705106), *PNAS* 106, 19780\u201319784. "
                   "Stromboli: [Valdivia et al. (2023)](https://link.springer.com/article/10.1007/s00410-023-02024-w), *Contrib. Mineral. Petrol.* 178, 45. "
                   "Tg(H\u2082O): [Langhammer et al. (2021)](https://agupubs.onlinelibrary.wiley.com/doi/full/10.1029/2021GC009918), *GGG* 22, e2021GC009918.")
